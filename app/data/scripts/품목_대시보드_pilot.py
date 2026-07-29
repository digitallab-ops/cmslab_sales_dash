"""
품목_대시보드_pilot.py  —  품목/SKU별 매출 대시보드 (파일럿)
기존 매출 Dashboard_vf.py 와 독립적으로 동작
"""

import sys, json, webbrowser
from pathlib import Path
import openpyxl
import xlrd

SCRIPT_DIR  = Path(__file__).parent
INPUT_DIR   = SCRIPT_DIR / 'Input'
OUTPUT_HTML = SCRIPT_DIR / '품목_대시보드.html'

# ── 채널 → 팀 ────────────────────────────────────────────────────
CHANNEL_TEAM = {
    '올리브영':'RBD1팀','군납':'RBD1팀','쿠팡':'RBD1팀',
    '네이버':'RBD1팀','버티컬몰':'RBD1팀','제휴몰':'RBD1팀',
    '임직원몰':'RBD1팀','기타':'RBD1팀',
    '면세':'RBD2팀','다이소':'RBD2팀','계열사':'RBD2팀','코스트코(한국)':'RBD2팀',
    '큐텐/라쿠텐':'일본사업팀','아마존_일본':'일본사업팀','해외_오프라인(일본)':'일본사업팀',
    '파트너사':'중국사업팀','중국법인수출':'중국사업팀',
    '티몰(중국)':'중국사업팀','해외_오프라인_대만':'중국사업팀',
    '해외_오프라인_미국':'Global사업팀','해외_오프라인_CIS':'Global사업팀',
    '해외_오프라인_유럽':'Global사업팀','해외_오프라인_동남아':'Global사업팀',
    '해외_오프라인_중동':'Global사업팀','해외_오프라인_기타':'Global사업팀',
    '자사몰-해외':'GEC팀','아마존_미국':'GEC팀','아마존_기타':'GEC팀',
    '틱톡':'GEC팀','역직구몰':'GEC팀','해외_온라인 기타':'GEC팀',
    '종합병원':'메디컬팀','클리닉':'메디컬팀','엑스퍼트몰':'메디컬팀','대리점':'메디컬팀',
}
SALE_CATS = {'본품','본품세트','벌크','덕용'}

TEAM_ORDER = ['전사','RBD1팀','RBD2팀','일본사업팀','중국사업팀',
              '동북아MC팀','Global사업팀','GEC팀','메디컬팀']


def resolve_team(channel, brand):
    if channel == '자사몰-국내':
        return '동북아MC팀' if brand in ('CFC','KTZ') else 'RBD1팀'
    return CHANNEL_TEAM.get(channel, '기타')


def find_latest(pattern):
    files = sorted(INPUT_DIR.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
    if not files:
        sys.exit(f'[오류] {pattern} 파일을 Input 폴더에서 찾을 수 없습니다.')
    return files[0]


# ── 데이터 로드 ───────────────────────────────────────────────────

def load_factory():
    path = find_latest('공장품목조회*.xls')
    print(f'  공장품목조회: {path.name}')
    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)
    hdr = [str(ws.cell_value(0, j)).strip() for j in range(ws.ncols)]
    ci = {c: hdr.index(c) for c in ['품목','품목명','제품군명','대분류명','품목분류','품목군']}
    # '테마' 열은 파일 구조 변경 대비 안전하게 조회
    theme_idx = hdr.index('테마') if '테마' in hdr else None
    items = {}
    for r in range(1, ws.nrows):
        code = str(ws.cell_value(r, ci['품목'])).strip()
        if not code:
            continue
        dl = str(ws.cell_value(r, ci['대분류명'])).strip()
        theme = ''
        if theme_idx is not None:
            theme = str(ws.cell_value(r, theme_idx)).strip()
        items[code] = {
            'sku_name':   str(ws.cell_value(r, ci['품목명'])).strip(),
            'brand':      str(ws.cell_value(r, ci['제품군명'])).strip(),
            'dl_cat':     dl,
            'item_cat':   str(ws.cell_value(r, ci['품목분류'])).strip(),
            'item_group': str(ws.cell_value(r, ci['품목군'])).strip(),
            'sale_type':  '판매품' if dl in SALE_CATS else '증정품',
            'theme':      theme,
        }
    return items


def load_channel_map():
    """거래처 → (채널, 국가, 거래처명) 매핑 로드.
    시트: '거래처 → 영업그룹 Mapping_로지'
    - B열(idx 1) 거래처코드, C열(idx 2) 거래처명, D열(idx 3) 채널(lv1), G열(idx 6) 국가
    """
    path = find_latest('1. 관리회계 기준정보*.xlsx')
    print(f'  기준정보: {path.name}')
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb['거래처 → 영업그룹 Mapping_로지']
    cmap, country_map, customer_map, in_data = {}, {}, {}, False
    for row in ws.iter_rows(values_only=True):
        if row[1] == '거래처코드':
            in_data = True; continue
        if not in_data or row[1] is None:
            continue
        try:
            code = str(int(float(row[1])))
        except:
            code = str(row[1]).strip()
        cust_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        lv1 = str(row[3]).strip() if row[3] else ''
        country = str(row[6]).strip() if len(row) > 6 and row[6] else ''
        if code and cust_name:
            customer_map[code] = cust_name
        if code and lv1:
            cmap[code] = lv1
        if code and country:
            country_map[code] = country
    wb.close()
    return cmap, country_map, customer_map


def load_channel_order():
    """팀참고 파일에서 채널 노출 순서 로드.
    시트 Sheet1의 첫 섹션(row 7~44 근처)에 B열=Code, C열=채널명이 있음.
    Code가 A로 시작하는 첫 연속 블록만 채널 순서로 사용.
    반환: [채널명, ...] (파일 등장 순서)
    """
    try:
        path = find_latest('팀참고*.xlsx')
    except SystemExit:
        print('  [경고] 팀참고 파일 없음 — 채널 순서는 알파벳 순으로 사용')
        return []
    print(f'  팀참고: {path.name}')
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb['Sheet1'] if 'Sheet1' in wb.sheetnames else wb.active
    order = []
    seen = set()
    for row in ws.iter_rows(values_only=True):
        code = row[1] if len(row) > 1 else None
        ch = row[2] if len(row) > 2 else None
        if isinstance(code, str) and code.startswith('A') and ch:
            ch_name = str(ch).strip()
            if ch_name and ch_name not in seen:
                order.append(ch_name)
                seen.add(ch_name)
    wb.close()
    return order


def load_sales(year, factory, cmap, country_map, customer_map):
    path = find_latest(f'통합매출_{year}*.xlsx')
    print(f'  통합매출 {year}: {path.name}')
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    recs = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        date_val = row[1]; cust_raw = row[2]; sku_raw = row[4]
        qty  = float(row[6] or 0)
        rev  = float(row[7] or 0)
        cost = float(row[8] or 0)
        if not (date_val and sku_raw and (rev or qty)):
            continue
        try:
            d = int(float(str(date_val))); month = (d % 10000) // 100
            if not 1 <= month <= 12: continue
        except:
            continue
        sku = str(sku_raw).strip()
        try:
            cust = str(int(float(str(cust_raw))))
        except:
            cust = str(cust_raw).strip()
        item     = factory.get(sku, {})
        channel  = cmap.get(cust, '')
        country  = country_map.get(cust, '')
        customer = customer_map.get(cust, '') or cust    # 거래처명 없으면 코드
        brand    = item.get('brand', '')
        recs.append({
            'yr':         year,
            'month':      month,
            'quarter':    f'Q{(month-1)//3+1}',
            'team':       resolve_team(channel, brand),
            'channel':    channel,
            'country':    country,
            'customer':   customer,
            'brand':      brand,
            'theme':      item.get('theme', ''),
            'dl_cat':     item.get('dl_cat', ''),
            'item_group': item.get('item_group', ''),
            'item_cat':   item.get('item_cat', ''),
            'sku':        sku,
            'sku_name':   item.get('sku_name', str(row[5] or '')),
            'sale_type':  item.get('sale_type', '증정품'),
            'qty':        qty,
            'rev':        rev,
            'cost':       cost,
        })
    wb.close()
    return recs


def aggregate(recs):
    try:
        import pandas as pd
    except ImportError:
        sys.exit('[오류] pandas 가 설치되어 있지 않습니다.')
    df = pd.DataFrame(recs)
    if df.empty:
        return []
    dims = ['yr','month','quarter','team','channel','country','customer','brand','theme',
            'dl_cat','item_group','item_cat','sku','sku_name','sale_type']
    agg = (df.groupby(dims, dropna=False)[['qty','rev','cost']]
             .sum().reset_index())
    agg['qty']  = agg['qty'].round(0).astype(int)
    agg['rev']  = agg['rev'].round(0).astype(int)
    agg['cost'] = agg['cost'].round(0).astype(int)
    return agg.to_dict('records')


# ── 메타 데이터 추출 (사이드바 트리 구성용) ─────────────────────────

def build_tree_meta(records):
    """품목군 → [품목분류] 매핑 (정렬된 dict)"""
    from collections import defaultdict
    tree = defaultdict(set)
    for r in records:
        tree[r['item_group']].add(r['item_cat'])
    return {g: sorted(cats) for g, cats in sorted(tree.items())}


# ── Chart.js 로컬 파일 로드 ───────────────────────────────────────

def load_chartjs() -> str:
    """chart.umd.js를 스크립트 폴더에서 탐색 후 내용 반환."""
    for name in ('chart.umd.js', 'chart.js'):
        p = SCRIPT_DIR / name
        if p.exists():
            return p.read_text(encoding='utf-8')
    return ''   # CDN fallback 으로 처리


# ── HTML 생성 ─────────────────────────────────────────────────────

CSS_ITEM = """
*{box-sizing:border-box;margin:0;padding:0}
:root{
  --blue:#1a56a0;--blue-dark:#154080;--bg:#f5f6fa;--card:#fff;
  --border:#e5e7eb;--text:#1a1a2e;--muted:#6c757d;
  --ff:'Malgun Gothic','Apple SD Gothic Neo',sans-serif;
  --radius:10px;--shadow:0 1px 4px rgba(0,0,0,.05);
  --pos-dark:#375623;--neg-dark:#C55A11;
  --c24:#6b7280;--c25:#5a96c8;--c26:#5C2508;
}
body{font-family:var(--ff);background:var(--bg);color:var(--text);font-size:13px;line-height:1.5;overflow-x:hidden}
/* STICKY */
.sticky-header   {position:sticky;top:0;z-index:200}
.sticky-filterbar{position:sticky;top:44px;z-index:190}
/* HEADER */
.header{background:#fff;height:44px;display:flex;align-items:center;padding:0 clamp(12px,2vw,24px);gap:14px;
  box-shadow:0 2px 8px rgba(0,0,0,.10);border-bottom:1px solid var(--border)}
.hdr-company{color:#833C0C;font-size:clamp(14px,1vw,15px);font-weight:800;white-space:nowrap}
.hdr-divider{width:1px;height:20px;background:#e5e7eb}
.hdr-right{margin-left:auto;color:var(--muted);font-size:11px;white-space:nowrap}
.hdr-right strong{color:var(--text);font-size:12px}
.hdr-tabs{display:flex;align-items:stretch;margin-left:12px;gap:0}
.hdr-tab{padding:0 16px;height:44px;display:flex;align-items:center;font-size:12px;font-weight:600;
  color:var(--muted);cursor:pointer;border:none;background:none;border-bottom:3px solid transparent;
  font-family:var(--ff);white-space:nowrap;transition:all .15s}
.hdr-tab:hover{color:var(--blue)}
.hdr-tab.active{color:var(--blue);border-bottom:3px solid var(--blue)}
/* 탭 패널 */
.tab-pane.hidden{display:none}
/* FILTERBAR */
.filterbar{background:var(--blue);padding:6px clamp(12px,2vw,24px) 7px;display:flex;flex-direction:column;gap:0;
  border-top:1px solid rgba(255,255,255,.12);box-shadow:0 3px 8px rgba(0,0,0,.15)}
/* overflow:visible로 두어 자식 요소(다중선택 드롭박스 메뉴)가 필터바 아래로 튀어나올 수 있게 함 */
.filter-row{display:flex;align-items:center;gap:24px;flex-wrap:nowrap;padding:3px 0;min-width:0;overflow:visible}
.fg{display:flex;align-items:center;gap:4px;flex-wrap:nowrap;flex-shrink:0}
.fg-label{display:inline-flex;align-items:center;justify-content:center;background:#DDEBF7;color:#1a56a0;
  font-size:11px;font-weight:800;white-space:nowrap;padding:2px 8px;border-radius:5px;letter-spacing:.3px;
  margin-right:4px;box-shadow:0 1px 3px rgba(0,0,0,.18);border:1px solid #9cc2e8}
.fg-sep{color:rgba(255,255,255,.30);padding:0 4px;font-size:13px;font-weight:300}
.bu-sep{padding:0 1px;margin:0 -3px}
.team-fg{flex-shrink:0}
.pb{padding:2px 10px;border-radius:20px;cursor:pointer;font-size:11px;font-family:var(--ff);
  white-space:nowrap;transition:background .14s,color .14s;line-height:1.7;font-weight:700;
  background:rgba(255,255,255,0.12);border:1px solid rgba(255,255,255,0.3);color:rgba(255,255,255,0.75)}
.pb:hover{background:rgba(255,255,255,.22);color:#fff}
.pb.active{background:#fff;border-color:#fff;color:var(--blue)}
.fg-label.single-mode{background:#f59e0b;border-color:#d97706;color:#fff}
.fg-label.toggleable{cursor:pointer;user-select:none;transition:background .15s,color .15s}
.fg-label.toggleable:hover{filter:brightness(0.95)}
.ch-select{height:24px;padding:0 24px 0 10px;border-radius:20px;border:1px solid rgba(255,255,255,0.3);
  background:rgba(255,255,255,0.12) url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='10' height='6'%3E%3Cpath d='M0 0l5 6 5-6z' fill='rgba(255,255,255,0.7)'/%3E%3C/svg%3E") no-repeat right 8px center;
  color:rgba(255,255,255,0.85);font-size:11px;font-family:var(--ff);font-weight:600;cursor:pointer;outline:none;
  min-width:100px;appearance:none;-webkit-appearance:none}
.ch-select:focus,.ch-select:hover{background-color:rgba(255,255,255,.22);color:#fff}
.ch-select option{background:#1a56a0;color:#fff;font-weight:600}
/* 다중선택 드롭박스 (브랜드/채널/국가) */
.ms-drop{position:relative;display:inline-block}
.ms-btn{height:24px;padding:0 22px 0 10px;border-radius:20px;border:1px solid rgba(255,255,255,0.3);
  background:rgba(255,255,255,0.12);color:rgba(255,255,255,0.85);
  font-size:11px;font-family:var(--ff);font-weight:600;cursor:pointer;outline:none;
  min-width:100px;max-width:180px;display:inline-flex;align-items:center;gap:4px;position:relative}
.ms-btn:hover{background:rgba(255,255,255,0.22);color:#fff}
.ms-drop.open .ms-btn{background:#fff;color:var(--blue);border-color:#fff}
.ms-label{flex:1;overflow:hidden;text-overflow:ellipsis;white-space:nowrap;text-align:left}
.ms-arrow{position:absolute;right:8px;top:50%;transform:translateY(-50%);font-size:8px;pointer-events:none;transition:transform .15s}
.ms-drop.open .ms-arrow{transform:translateY(-50%) rotate(180deg)}
.ms-menu{display:none;position:absolute;top:calc(100% + 3px);left:0;min-width:100%;max-height:280px;
  background:#fff;border:1px solid var(--border);border-radius:6px;
  box-shadow:0 4px 12px rgba(0,0,0,.15);overflow-y:auto;z-index:210;padding:4px 0}
.ms-drop.open .ms-menu{display:block}
.ms-item{display:flex;align-items:center;gap:6px;padding:5px 10px;cursor:pointer;
  font-size:11px;color:#374151;white-space:nowrap;user-select:none}
.ms-item:hover{background:#f0f6fc}
.ms-item input[type=checkbox]{width:13px;height:13px;accent-color:var(--blue);cursor:pointer;flex-shrink:0}
.ms-item span{overflow:hidden;text-overflow:ellipsis}
.ms-item-all{border-bottom:1px solid var(--border);font-weight:700;color:var(--blue)}
/* LAYOUT */
.layout{display:flex;min-height:calc(100vh - var(--sticky-total,170px));padding-left:clamp(12px,2vw,24px)}
/* SIDEBAR */
.sidebar{width:280px;flex-shrink:0;background:#fff;border-right:1px solid var(--border);
  overflow-y:auto;position:sticky;top:var(--sticky-total,170px);max-height:calc(100vh - var(--sticky-total,170px));font-size:12px;
  z-index:155}  /* overlay(150) 위, expanded chart(160) 아래 — 확대 상태에서도 편집 가능 */
.sidebar-hdr{padding:8px 10px;background:#f8fafc;border-bottom:1px solid var(--border);
  display:flex;align-items:center;justify-content:space-between;font-weight:700;font-size:12px;color:#374151;
  position:sticky;top:0;z-index:10}
/* 사이드바 헤더: 제목이 좌측에 수직 중앙 정렬, 우측 2줄 버튼 세트 */
.sidebar-hdr-2rows{display:flex;align-items:center;gap:8px;padding:6px 10px}
.sidebar-hdr-2rows .sb-title{font-weight:700;color:#374151;flex:1;font-size:12px}
.sidebar-hdr-2rows .sb-btns-col{display:flex;flex-direction:column;gap:3px}
.sidebar-hdr-2rows .sb-btns-row{display:flex;gap:3px}
.sb-btn{font-size:10px;padding:2px 6px;border:1px solid var(--border);border-radius:4px;cursor:pointer;
  background:#fff;color:var(--muted);font-family:var(--ff);white-space:nowrap;text-align:center;box-sizing:border-box}
.sb-btn:hover{background:var(--blue);color:#fff;border-color:var(--blue)}
/* 전체=펼치기, 해제=닫기 폭 완전 통일 (min-width→width로 픽셀 고정) */
.sb-btn-a{width:64px}   /* 전체 · ▾ 펼치기 */
.sb-btn-b{width:54px}   /* 해제 · ▸ 닫기 */
/* 품목군(그룹) 헤더 — 파랑 톤 배경으로 품목분류와 명확히 구분 */
.tree-grp-hdr{display:flex;align-items:center;gap:8px;padding:5px 10px;cursor:pointer;
  background:#DDEBF7;border-top:1px solid #b8cee5;border-bottom:1px solid #b8cee5;
  user-select:none}
.tree-grp-hdr:hover{background:#c9dcef}
.tree-grp-cb{accent-color:var(--blue);width:14px;height:14px;flex-shrink:0;cursor:pointer}
.tree-grp-lbl{font-weight:700;color:#1e3a5f;font-size:12px;flex:1}
/* 토글 화살표는 이전 크기 유지 (터치/클릭 편의) */
.tree-arr{font-size:26px;color:var(--muted);transition:transform .2s,background .12s;flex-shrink:0;
  cursor:pointer;padding:11px 15px;margin:-11px -15px -11px 0;border-radius:6px;line-height:1}
.tree-arr:hover{background:rgba(0,0,0,.08);color:var(--blue)}
.tree-arr.closed{transform:rotate(-90deg)}
.tree-cats{overflow:hidden;transition:max-height .25s ease;max-height:2000px}
.tree-cats.closed{max-height:0}
/* 품목분류(카테고리) 아이템 — 흰 배경 + 깊은 들여쓰기로 위계 표현 */
.tree-cat-item{display:flex;align-items:center;gap:6px;padding:4px 8px 4px 24px;
  background:#fff;border-bottom:1px solid #eef1f4}
.tree-cat-item:hover{background:#f0f6fc}
.tree-cat-cb{accent-color:var(--blue);width:13px;height:13px;flex-shrink:0;cursor:pointer}
.tree-cat-item label{font-size:12px;color:#374151;cursor:pointer;flex:1;
  overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
/* 품목분류 우측 SKU 편집 버튼 */
.tree-sku-btn{flex-shrink:0;width:20px;height:20px;padding:0;border:1px solid var(--border);
  background:#f8fafc;color:var(--muted);border-radius:4px;cursor:pointer;font-size:11px;
  display:flex;align-items:center;justify-content:center;font-family:var(--ff)}
.tree-sku-btn:hover{background:var(--blue);color:#fff;border-color:var(--blue)}
/* 품목분류에 SKU가 일부만 선택된 경우 — 라벨을 주황색으로 강조 */
.tree-cat-item.partial-sku label{color:#d97706;font-weight:600}
.tree-cat-item.partial-sku label::after{content:' (일부)';font-size:10px;color:#d97706;font-weight:500}

/* SKU 선택 모달 */
.sku-modal-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,.4);z-index:250;animation:fadeIn .18s ease}
.sku-modal-bg.open{display:block}
.sku-modal{display:none;position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);
  background:#fff;width:560px;max-width:92vw;max-height:82vh;
  border-radius:10px;box-shadow:0 8px 40px rgba(0,0,0,.28);z-index:260;
  flex-direction:column;overflow:hidden}
.sku-modal.open{display:flex}
.sku-modal-hdr{padding:12px 16px;background:var(--blue-dark);color:#fff;
  display:flex;align-items:center;gap:10px;flex-shrink:0}
.sku-modal-hdr > span{flex:1;font-size:14px;font-weight:700;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
.sku-modal-close{background:transparent;border:none;color:#fff;cursor:pointer;
  font-size:18px;padding:0 4px;font-family:var(--ff)}
.sku-modal-close:hover{opacity:0.7}
.sku-modal-toolbar{padding:8px 16px;display:flex;align-items:center;gap:8px;
  border-bottom:1px solid var(--border);background:#f8fafc;flex-shrink:0}
.sku-modal-count{margin-left:auto;font-size:11px;color:var(--muted);font-weight:600}
.sku-modal-body{flex:1;overflow-y:auto;padding:4px 12px;min-height:0}
.sku-item{display:flex;align-items:center;gap:8px;padding:6px 8px;border-bottom:1px solid #f3f4f6}
.sku-item:hover{background:#f0f6fc}
.sku-item input[type=checkbox]{width:14px;height:14px;accent-color:var(--blue);cursor:pointer;flex-shrink:0}
.sku-item label{font-size:12px;color:#374151;cursor:pointer;flex:1;
  overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.sku-item .sku-code{font-family:'Consolas',monospace;color:var(--blue);font-weight:600;margin-right:8px}
/* 판매품/증정품 구분 배지 (SKU 모달 및 품목별 매출 표에서 공용)
   SKU# 앞에 위치하므로 좌우 margin 대신 우측 margin으로 여백 확보 */
.sale-badge{display:inline-block;padding:1px 6px;border-radius:4px;
  font-size:10px;font-weight:700;line-height:1.3;margin-right:6px;flex-shrink:0;
  white-space:nowrap;vertical-align:middle;font-family:var(--ff)}
.sb-sale{background:#dcfce7;color:#166534;border:1px solid #86efac}
.sb-gift{background:#fef3c7;color:#92400e;border:1px solid #fcd34d}
.sb-none{background:#f3f4f6;color:#6b7280;border:1px solid #d1d5db}
/* CONTENT */
.content{flex:1;min-width:0;padding:10px 16px;display:flex;flex-direction:column;gap:10px;padding-bottom:40px}
/* CHART */
.chart-card{background:var(--card);border:1px solid var(--border);border-radius:var(--radius);box-shadow:var(--shadow);overflow:hidden}
.chart-card-hdr{padding:8px 14px;border-bottom:1px solid var(--border);display:flex;align-items:center;gap:10px}
.chart-card-title{font-size:13px;font-weight:700;color:var(--text);flex:1}
.chart-card-title.title-rev{color:var(--blue)}   /* 매출액 계열 — 파랑 */
.chart-card-title.title-qty{color:#166534}       /* 매출수량 계열 — 진녹색 */
.chart-card-body{padding:10px 14px}
.charts-full{display:flex;flex-direction:column;gap:10px}
/* 누적 매출액 45% : 누적 매출수량 55% (매출수량 그래프 폭 확대) */
.charts-half{display:grid;grid-template-columns:0.9fr 1.1fr;gap:10px}
/* 차트 확대 (매출 Dashboard_vf.py와 동일 로직) */
.chart-expand-btn{display:flex;align-items:center;justify-content:center;width:24px;height:24px;
  border-radius:5px;border:1px solid var(--border);background:#fff;cursor:pointer;color:var(--muted);
  font-size:13px;flex-shrink:0;transition:all .14s}
.chart-expand-btn:hover{background:#f0f4ff;color:var(--blue);border-color:var(--blue)}
.chart-expanded{position:fixed!important;z-index:160;background:#fff;border-radius:var(--radius);
  box-shadow:0 8px 40px rgba(0,0,0,.22);display:flex;flex-direction:column;
  transition:top .25s ease,left .25s ease,width .25s ease,height .25s ease}
.chart-expanded .chart-card-body{flex:1;height:auto!important;min-height:0}
.chart-expanded .stbl-toggle,.chart-expanded .stbl-wrap{flex-shrink:0}
.chart-expanded .stbl-wrap.open{max-height:220px}
.chart-overlay-bg{position:fixed;inset:0;background:rgba(0,0,0,.25);z-index:150;cursor:pointer;animation:fadeIn .2s ease}
@keyframes fadeIn{from{opacity:0}to{opacity:1}}
/* TABLE
   sticky 규칙: .tbl-toolbar와 .tbl-outer thead를 별도로 sticky 처리
   - 툴바(.tbl-toolbar): sticky top=var(--sticky-total) — KPI 카드 바로 밑에 고정
   - 컬럼 헤더(thead): sticky top=var(--sticky-total)+툴바높이(--tbl-toolbar-h) — 툴바 바로 밑에 고정
   - .tbl-card의 overflow는 clip 사용 (scroll container 생성하지 않아 sticky 자식이 뷰포트에 정상 앵커링됨) */
.tbl-card{background:var(--card);border:1px solid var(--border);border-radius:var(--radius);box-shadow:var(--shadow);overflow:visible}
.tbl-toolbar{display:flex;align-items:center;padding:8px 12px 10px;border-bottom:1px solid var(--border);gap:8px;
  background:#fff;position:sticky;top:var(--sticky-total,170px);z-index:6;
  border-top-left-radius:var(--radius);border-top-right-radius:var(--radius);
  box-shadow:0 4px 0 #fff}  /* 툴바 하단 4px 흰색 그림자로 thead와의 서브픽셀 gap 커버 */
.tbl-title{font-size:13px;font-weight:700;color:var(--text);flex:1}
.tbl-count{font-size:11px;color:var(--muted);white-space:nowrap}
.tbl-search{border:1px solid var(--border);border-radius:6px;padding:4px 10px;font-size:11px;font-family:var(--ff);outline:none;width:180px}
.tbl-expand-btn{font-size:11px;padding:3px 8px;background:#f0f4f8;color:#374151;border:1px solid var(--border);border-radius:4px;cursor:pointer;font-family:var(--ff);white-space:nowrap}
.tbl-expand-btn:hover{background:var(--blue);color:#fff;border-color:var(--blue)}
.tbl-search:focus{border-color:var(--blue)}
/* .tbl-outer overflow는 visible 유지 → thead의 position:sticky가 뷰포트 앵커링됨.
   두 표(품목별/채널별)가 같은 % colgroup 폭을 쓰므로 컨테이너 폭이 같은 한 시각적 정렬이 완전 일치.
   라벨 열은 40% 배정하여 계층 깊은 이름(▸ SKU 등)도 잘 보이도록 넓게 확보. */
.tbl-outer{overflow:visible}
table{border-collapse:collapse;white-space:nowrap;font-size:12px;table-layout:fixed;width:100%}
/* 라벨 열: 컨텐츠에 맞춰 ellipsis */
td.fix-col,th.fix-col{overflow:hidden;text-overflow:ellipsis}
/* thead top offset을 툴바 하단에 2px 오버랩시켜 sticky 요소 간 gap 완전 제거 */
.tbl-outer thead{position:sticky;top:calc(var(--sticky-total,170px) + var(--tbl-toolbar-h,42px) - 2px);z-index:5}
thead tr th{padding:6px 10px;background:#1F4E78;color:#fff;font-size:11px;font-weight:700;
  border:1px solid #3d5a80;text-align:center;white-space:nowrap}
thead tr.sh2 th{background:#374151;font-size:11px}
/* 연도별 헤더 밴드 (row 1 — 연도명 배경 = 연도 계열 컬러) */
thead th.yh-2024{background:#6b7280}                    /* 24년: 중립 그레이 */
thead th.yh-2025{background:var(--c25);color:#fff}      /* 25년: 파랑 */
thead th.yh-2025{background:#7ba9d1;color:#fff}          /* 25년: 파랑 (톤다운 2단계) */
thead th.yh-2026{background:#c88870;color:#fff}          /* 26년: 적갈 (톤다운 2단계) */
thead th.yh-diff {background:#7a848e;color:#fff}         /* 증감: 중립 그레이 (덜 진한) */
/* 연도별 서브 헤더 (row 2 — 지표명) — 조금 어둡게 (톤다운) */
thead th.ys-2024{background:#8a8f96}
thead th.ys-2025{background:#5f8ab4}                     /* 파랑 어둡게 (덜 진하게) */
thead th.ys-2026{background:#a56e57}                     /* 적 어둡게 (덜 진하게) */
thead th.ys-diff {background:#576068;color:#fff}         /* 증감 서브헤더 */
/* 지표명 우측 단위 (개/백만원/%/%p) — 다음 줄에 작게 표시하여 헤더 폭 절약 */
thead th .unit{display:block;font-size:10px;font-weight:500;opacity:.9;margin-top:1px;line-height:1.1}
/* 계층별 데이터 셀 배경은 아래 tr.grp-row/cat-row/sku-row/chan-row 규칙에서 정의 (계층별 그라디언트) */
/* 연도 그룹 시작 위치에 좌측 구분선 */
thead th.y-sep-l,tbody td.y-sep-l{border-left:2px solid #b8cee5}
th.fix-col{position:sticky;left:0;z-index:6;background:#1F4E78}
tbody tr{transition:background .08s}
tbody tr:nth-child(even){background:#fafafa}
tbody tr:hover{background:#f0f4ff}
tbody td{padding:5px 10px;border-bottom:1px solid #f3f4f6;border-right:1px solid #f3f4f6;
  text-align:right;font-size:12px;white-space:nowrap}
td.fix-col{position:sticky;left:0;background:#fff;z-index:3;text-align:left;border-right:2px solid #e5e7eb;
  min-width:160px;max-width:240px;overflow:hidden;text-overflow:ellipsis}
tr:nth-child(even) td.fix-col{background:#fafafa}
tr:hover td.fix-col{background:#f0f4ff}
/* ─── 계층별 3색 팔레트 (품목군 > 품목분류 > SKU > 채널) ───
   fix-col(라벨 열): 파랑 그라디언트 (진→최옅)
   2025 데이터 열:  녹색 그라디언트 (진→최옅)
   2026 데이터 열:  붉은색 그라디언트 (진→최옅)
   계층이 깊어질수록 배경이 옅어져 위계가 자연스럽게 드러남 */

/* 품목군 (최상위) — 톤다운 2단계 */
tr.grp-row td{font-weight:800;color:#1e3a5f;font-size:13px;border-top:2px solid #9ab5cb}
tr.grp-row td.fix-col{background:#c8d9e6!important;padding-left:8px}                     /* 파랑 진 (덜 진하게) */
tr.grp-row td.yc-2025{background:#bfd6e8!important;color:#1e3a5f}                        /* 파랑 진 (덜 진하게) */
tr.grp-row td.yc-2026{background:#e6b5a3!important;color:#5a2312}                        /* 적색 진 (덜 진하게) */
tr.grp-row td.yc-2024{background:#c8c8c8!important;color:#222}                           /* 회색 진 */

/* 품목분류 (2단계) — 톤다운 2단계 */
tr.cat-row td{font-weight:700;color:#1e3a5f;font-size:12px;border-top:1px solid #d5e2ec}
tr.cat-row td.fix-col{background:#dde8f0!important;padding-left:18px}                    /* 파랑 중 (덜 진) */
tr.cat-row td.yc-2025{background:#d8e5ef!important;color:#1e3a5f}                        /* 파랑 중 (덜 진) */
tr.cat-row td.yc-2026{background:#f0d3c8!important;color:#5a2312}                        /* 적색 중 (덜 진) */
tr.cat-row td.yc-2024{background:#dcdcdc!important;color:#333}                           /* 회색 중 */

/* SKU (3단계) — 톤다운 2단계 */
tr.sku-row td{border-top:1px solid #e5e9ee;font-weight:600}
tr.sku-row td.fix-col{background:#f0f4f8;padding-left:28px;color:#1a1a2e;font-size:12px}  /* 파랑 옅 */
tr.sku-row td.yc-2025{background:#ecf1f6}                                                 /* 파랑 옅 */
tr.sku-row td.yc-2026{background:#f7e8df}                                                 /* 적색 옅 */
tr.sku-row td.yc-2024{background:#f0f0f0}                                                 /* 회색 옅 */

/* 채널 (4단계 = 최하위) — 이미 옅어 미세 조정만 */
tr.chan-row td{color:#4b5563;font-size:11px;border-top:1px dashed #e5e9ee}
tr.chan-row td.fix-col{background:#f9fbfd!important;padding-left:52px;font-weight:500;color:#4b5563}
tr.chan-row td.yc-2025{background:#f4f8fc!important;color:#1e3a5f}
tr.chan-row td.yc-2026{background:#fdf5ef!important;color:#5a2312}
tr.chan-row td.yc-2024{background:#f9f9f9!important;color:#4b5563}

/* 짝수 행 nth-child 오버라이드는 계층 색을 우선하도록 무효화 */
tbody tr:nth-child(even) td{background-color:inherit}

/* 접힘 상태 */
tr.sku-row.hidden{display:none}
tr.cat-row.hidden{display:none}
tr.chan-row.hidden{display:none}
/* 합계 행 (두 표 공통): 상단 구분선, 톤다운 남색, 두꺼운 볼드
   '합계' 라벨(fix-col)만 가운데 정렬, 숫자 셀은 다른 행처럼 우측 정렬 유지 */
tr.total-row td{background:#5b7ba3!important;color:#fff!important;font-weight:800;font-size:12.5px;
  border-top:3px solid #3d5a80;padding-top:6px;padding-bottom:6px;text-align:right}
tr.total-row td.fix-col{background:#5b7ba3!important;color:#fff!important;
  padding-left:0!important;text-align:center!important}
tr.total-row td.yc-2025,tr.total-row td.yc-2026,tr.total-row td.yc-2024,tr.total-row td.yc-diff{
  background:#5b7ba3!important;color:#fff!important;text-align:right}
/* hover 상태에서 미세한 밝기 유지 */
tr.chan-row:hover td.fix-col{filter:brightness(0.96)}
tr.sku-row:hover td.fix-col{filter:brightness(0.98)}

/* ── 증감 컬럼: 계층별 중립 그레이 그라디언트 + 양수/음수 색상 */
tr.grp-row  td.yc-diff{background:#d5dde3!important}
tr.cat-row  td.yc-diff{background:#e2e8ed!important}
tr.sku-row  td.yc-diff{background:#eef2f5}
tr.chan-row td.yc-diff{background:#f5f7f9!important}
tr.n2-ch    td.yc-diff{background:#d5dde3!important}
tr.n2-cs    td.yc-diff{background:#dde3e8!important}
tr.n2-gp    td.yc-diff{background:#e6ebf0!important}
tr.n2-ct    td.yc-diff{background:#eef2f5!important}
tr.n2-sk    td.yc-diff{background:#f5f7f9!important}
/* 증감 값 색상 (양수/음수/중립) */
td.yc-diff.diff-pos{color:#166534;font-weight:700}
td.yc-diff.diff-neg{color:#b91c1c;font-weight:700}
td.yc-diff.diff-neu{color:#9ca3af}

/* ─── 두 번째 표: 채널 > 거래처 > 품목군 > 품목분류 > SKU (5단계) ───
   fix-col 파랑 그라디언트 5단계 (진→최옅), 데이터 열도 25/26 각각 5단계 */
/* 레벨 1: 채널 (최상위) — 톤다운 */
tr.n2-ch td{font-weight:800;color:#1e3a5f;font-size:13px;border-top:2px solid #9ab5cb}
tr.n2-ch td.fix-col{background:#c8d9e6!important;padding-left:8px}
tr.n2-ch td.yc-2025{background:#bfd6e8!important;color:#1e3a5f}
tr.n2-ch td.yc-2026{background:#e6b5a3!important;color:#5a2312}
tr.n2-ch td.yc-2024{background:#c8c8c8!important;color:#222}
/* 레벨 2: 거래처 — 톤다운 */
tr.n2-cs td{font-weight:700;color:#1e3a5f;font-size:12px;border-top:1px solid #c8d9e5}
tr.n2-cs td.fix-col{background:#d4e1ec!important;padding-left:20px}
tr.n2-cs td.yc-2025{background:#cddceb!important;color:#1e3a5f}
tr.n2-cs td.yc-2026{background:#edc4b3!important;color:#5a2312}
tr.n2-cs td.yc-2024{background:#d0d0d0!important;color:#222}
/* 레벨 3: 품목군 — 톤다운 */
tr.n2-gp td{font-weight:700;color:#1e3a5f;font-size:12px;border-top:1px solid #d5e2ec}
tr.n2-gp td.fix-col{background:#dfe9f1!important;padding-left:32px}
tr.n2-gp td.yc-2025{background:#dbe6f0!important;color:#1e3a5f}
tr.n2-gp td.yc-2026{background:#f0d3c8!important;color:#5a2312}
tr.n2-gp td.yc-2024{background:#dcdcdc!important;color:#333}
/* 레벨 4: 품목분류 — 톤다운 */
tr.n2-ct td{font-weight:600;color:#374151;font-size:11.5px;border-top:1px solid #e0e6ed}
tr.n2-ct td.fix-col{background:#ecf0f5!important;padding-left:44px}
tr.n2-ct td.yc-2025{background:#e8eff5!important;color:#1e3a5f}
tr.n2-ct td.yc-2026{background:#f3ddd0!important;color:#5a2312}
tr.n2-ct td.yc-2024{background:#e5e5e5!important;color:#374151}
/* 레벨 5: SKU (최하위) — 이미 옅어 미세 조정만 */
tr.n2-sk td{color:#4b5563;font-size:11px;border-top:1px dashed #e5e9ee;font-weight:500}
tr.n2-sk td.fix-col{background:#f9fbfd!important;padding-left:56px;font-weight:500;color:#4b5563}
tr.n2-sk td.yc-2025{background:#f4f8fc!important;color:#1e3a5f}
tr.n2-sk td.yc-2026{background:#fdf5ef!important;color:#5a2312}
tr.n2-sk td.yc-2024{background:#f9f9f9!important;color:#4b5563}
/* 접힘 상태 (기본 데이터 로드 후 채널 이하만 노출) */
tr.n2-cs.hidden,tr.n2-gp.hidden,tr.n2-ct.hidden,tr.n2-sk.hidden{display:none}
tr.n2-sk:hover td.fix-col{filter:brightness(0.98)}
/* 표 내부 토글 화살표: 사이드바 화살표 수준으로 크게 (클릭/터치 편의).
   font-size는 크게 하되 padding/margin trick으로 행 높이는 크게 늘어나지 않도록 함.
   행 내 라벨 텍스트 폰트는 별도 규칙으로 유지됨. */
.toggle-arrow{display:inline-block;cursor:pointer;font-size:22px;color:var(--muted);
  line-height:1;vertical-align:middle;padding:4px 8px;margin:-4px 2px -4px -4px;
  border-radius:5px;transition:background .12s,color .12s;font-weight:500}
.toggle-arrow:hover{background:rgba(0,0,0,.08);color:var(--blue)}
.neg{color:#dc2626}
.pos{color:#166534}
/* KPI STRIP — 4개 그룹(헤더바 + 3카드) 구조
   그룹 폭은 컨텐츠 성격에 맞게 차등: 매출/판매수량은 넉넉히, 품목정보는 축소 */
.item-kpi-strip{background:var(--bg);padding:8px clamp(12px,2vw,24px);
  display:grid;grid-template-columns:1.05fr 0.5fr 0.9fr 1.05fr 0.7fr;gap:10px;
  border-bottom:1px solid var(--border);position:sticky;top:var(--sticky-hf,114px);z-index:170}
.kpi-group.kpi-single .kpi-group-body{grid-template-columns:1fr}
.kpi-group{display:flex;flex-direction:column;border-radius:12px;overflow:hidden;
  box-shadow:var(--shadow);min-width:0}
/* 그룹 헤더바 — 진한 파랑 (선택탭/필터바 계열), 카드 높이의 약 절반 */
/* KPI 카드 톤다운 팔레트: 진파랑 → 부드러운 중간 파랑 */
.kpi-group-hdr{background:#2c5b93;color:#fff;
  padding:7px 10px;font-size:13px;font-weight:800;
  text-align:center;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;
  line-height:1.2}
.kpi-group-hdr span{font-size:10.5px;font-weight:500;opacity:.85;margin-left:3px}
.kpi-group-body{display:grid;grid-template-columns:repeat(3,1fr);gap:3px;
  background:#2c5b93;padding:3px}
.kpi-group.has-diff .kpi-group-body{grid-template-columns:1fr 1fr 1.5fr}
/* 개별 KPI 카드 — 필터바보다 조금 옅은 파랑 */
.kpi-card{background:#4a7fb8;color:#fff;
  padding:9px 4px;min-width:0;
  display:flex;flex-direction:column;align-items:center;justify-content:center;
  gap:3px;text-align:center;border-radius:6px}
.kpi-card .kpi-t{font-size:11px;font-weight:600;line-height:1.15;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:100%;opacity:.92}
.kpi-card .kpi-v{font-size:15px;font-weight:800;line-height:1.15;
  white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:100%}
/* 증감 카드 — 값 텍스트에만 부호 색상 반영 (파랑 배경 위에서 잘 보이도록 대비색) */
.kpi-card.kpi-diff.pos .kpi-v{color:#a8ecb0}
.kpi-card.kpi-diff.neg .kpi-v{color:#ffc0b3}
.kpi-card.kpi-diff.neu .kpi-v{color:#e5e7eb}
/* CHART LEGEND (title bar) */
.chart-legend{display:flex;align-items:center;gap:10px;margin-left:auto}
.cl-item{display:flex;align-items:center;gap:5px;font-size:11px;color:var(--text);font-weight:600;white-space:nowrap}
.cl-line{width:22px;height:3px;border-radius:2px;display:inline-block;vertical-align:middle}
/* STBL (차트 하단 토글 테이블) */
.stbl-toggle{display:flex;align-items:center;justify-content:center;gap:6px;width:100%;
  padding:5px 0;font-size:11px;font-weight:600;color:var(--muted);
  background:#fafafa;border:none;border-top:1px solid var(--border);
  cursor:pointer;transition:background .14s,color .14s;font-family:var(--ff)}
.stbl-toggle:hover{background:#f0f4ff;color:var(--blue)}
.stbl-arr{font-size:10px;transition:transform .3s;display:inline-block}
.stbl-toggle.open .stbl-arr{transform:rotate(180deg)}
.stbl-wrap{max-height:0;overflow:hidden;transition:max-height .35s ease,opacity .25s ease;opacity:0}
.stbl-wrap.open{max-height:200px;opacity:1;overflow-x:auto}
.stbl{width:100%;border-collapse:collapse;white-space:nowrap}
.stbl th,.stbl td{padding:4px 8px;text-align:right;font-size:10.5px;border-right:1px solid var(--border)}
/* 헤더행(구분/1월/2월/…)은 모두 가운데 정렬 */
.stbl th{text-align:center;font-weight:700;color:var(--muted);background:#f8fafc;border-bottom:1px solid var(--border)}
/* 데이터 행의 첫 열(연도명·채널명)은 좌측 정렬 유지, 헤더 첫 열은 위 규칙에 의해 가운데 유지 */
.stbl td:first-child{text-align:left;min-width:60px;background:#f8fafc;position:sticky;left:0;z-index:1}
.stbl th:first-child{min-width:60px;background:#f8fafc;position:sticky;left:0;z-index:1}
.stbl tr td{border-bottom:1px solid #f3f4f6}
.stbl .c24{color:var(--c24);font-weight:600}
.stbl .c25{color:var(--c25);font-weight:600}
.stbl .c26{color:var(--c26);font-weight:700}

/* ─── 반응형 미디어 쿼리 (매출 Dashboard_vf.py와 동일한 브레이크포인트) ─── */
@media (max-width:1280px) {
  .item-kpi-strip{padding:6px 16px;gap:10px}
  .kpi-card{padding:7px 3px}
  .kpi-card .kpi-v{font-size:13.5px}
  .kpi-group-hdr{font-size:12px;padding:6px 8px}
  .sticky-filterbar{top:44px}
  .filterbar{padding:6px 16px 7px}
  .header{padding:0 16px}
  .layout{padding-left:16px}
}
@media (max-width:1024px) {
  .item-kpi-strip{padding:6px 12px;gap:8px}
  .kpi-card{padding:6px 2px}
  .kpi-card .kpi-t{font-size:10px}
  .kpi-card .kpi-v{font-size:12px}
  .kpi-group-hdr{font-size:11px;padding:5px 6px}
  .layout{padding-left:12px}
  .content{padding:10px 12px}
}
"""


ITEM_GROUP_PREFIX = ['썬', '쿨링', '비비', 'MD']
ITEM_GROUP_SUFFIX = ['종료', '공란']


def sort_item_groups(groups):
    """품목군 정렬: 앞(썬/쿨링/비비/MD) → 중간(가나다순) → 뒤(종료/공란)"""
    grp_set = set(groups)
    prefix = [g for g in ITEM_GROUP_PREFIX if g in grp_set]
    suffix = [g for g in ITEM_GROUP_SUFFIX if g in grp_set]
    used   = set(prefix + suffix)
    middle = sorted(g for g in grp_set if g not in used)
    return prefix + middle + suffix


def build_sku_map(records, ordered_groups, tree_meta):
    """(gi, ci) → [{sku, sku_name, sale_type}] 매핑 — SKU 선택 모달용
    sale_type: '판매품' 또는 '증정품' (마스터 기준)"""
    sku_map = {}
    for gi, grp in enumerate(ordered_groups):
        cats = tree_meta[grp]
        for ci, cat in enumerate(cats):
            skus_seen = {}
            for r in records:
                if r.get('item_group') == grp and r.get('item_cat') == cat:
                    sku = r.get('sku', '')
                    if sku and sku not in skus_seen:
                        skus_seen[sku] = (r.get('sku_name', ''),
                                          r.get('sale_type', ''))
            key = f"{gi}_{ci}"
            sku_map[key] = [{'sku': s, 'sku_name': n, 'sale_type': st}
                            for s, (n, st) in sorted(skus_seen.items())]
    return sku_map


def make_html(records, chartjs_src, base_date='', channel_order=None, data_url=None):
    channel_order = channel_order or []
    channel_order_json = json.dumps(channel_order, ensure_ascii=False, separators=(',', ':'))
    tree_meta = build_tree_meta(records)
    # 브랜드 고정 순서 (JS의 BRAND_ORDER와 동일)
    BRAND_ORDER = ['CFC','CEX','DMB','SUS','PIN','KTZ','ETC']
    brand_set = {r['brand'] for r in records if r['brand']}
    brands = [b for b in BRAND_ORDER if b in brand_set] + sorted(b for b in brand_set if b not in BRAND_ORDER)
    countries = sorted({r['country'] for r in records if r.get('country')})
    all_channels = sorted({r['channel'] for r in records if r['channel']})

    tree_json = json.dumps(tree_meta, ensure_ascii=False)

    # 데이터 전달 방식: data_url이 있으면 RAW를 비우고 그 URL에서 fetch(온디맨드),
    # 없으면 기존처럼 RAW를 HTML에 임베드. (계산 로직은 어느 쪽이든 동일)
    if data_url:
        raw_decl = "let RAW = [];"
        init_js = (
            "fetch(" + json.dumps(data_url) + " + location.search)"
            ".then(function(r){ if(!r.ok) throw new Error('HTTP '+r.status); return r.json(); })"
            ".then(function(d){ RAW.length=0; for(var i=0;i<d.length;i++) RAW.push(d[i]);"
            " refreshAllDropdowns(); renderNow(); alignQuarterToYear(); })"
            ".catch(function(e){ document.body.insertAdjacentHTML('afterbegin',"
            "'<div style=\"padding:24px;font-family:sans-serif;color:#b91c1c\">데이터 로딩 실패: '+e+'</div>'); });"
        )
    else:
        data_json = json.dumps(records, ensure_ascii=False, separators=(',', ':'))
        raw_decl = "const RAW  = " + data_json + ";"
        init_js = "refreshAllDropdowns();\n  renderNow();\n  alignQuarterToYear();"

    if chartjs_src:
        chartjs_tag = f'<script>{chartjs_src}</script>'
    else:
        chartjs_tag = '<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>'

    # 복수선택 드롭박스용 옵션 HTML 생성 (전체 + 각 항목 체크박스)
    def _ms_items(vals):
        # 각 값에 checkbox 라벨. 기본 언체크(전체가 체크되어 있음)
        return '\n        '.join(
            f'<label class="ms-item"><input type="checkbox" value="{v}" onchange="onMsChange(this)"><span>{v}</span></label>'
            for v in vals)

    customers = sorted({r['customer'] for r in records if r.get('customer')})
    themes    = sorted({r['theme']    for r in records if r.get('theme')})

    brand_items    = _ms_items(brands)
    country_items  = _ms_items(countries)
    customer_items = _ms_items(customers)
    theme_items    = _ms_items(themes)

    # 품목군별 25년 매출액 합산 → 내림차순 정렬 → 스킨부스터 위치 조정
    grp_rev25: dict = {}
    for r in records:
        if r.get('yr') == 2025:
            g = r.get('item_group') or '(미분류)'
            grp_rev25[g] = grp_rev25.get(g, 0) + r.get('rev', 0)
    ordered_groups = sorted(tree_meta.keys(), key=lambda g: -grp_rev25.get(g, 0))
    # 스킨부스터를 바디케어 바로 다음으로 이동
    if '스킨부스터' in ordered_groups and '바디케어' in ordered_groups:
        ordered_groups = [g for g in ordered_groups if g != '스킨부스터']
        bi = ordered_groups.index('바디케어')
        ordered_groups.insert(bi + 1, '스킨부스터')
    grp_order_json = json.dumps(ordered_groups, ensure_ascii=False)

    # SKU 맵 생성 (품목분류별 SKU 리스트) — SKU 선택 모달용
    sku_map = build_sku_map(records, ordered_groups, tree_meta)
    sku_map_json = json.dumps(sku_map, ensure_ascii=False, separators=(',', ':'))

    # 사이드바 트리 HTML — 정렬 적용, 기본 접힘, 화살표만 토글
    sb_parts = []
    for gi, grp in enumerate(ordered_groups):
        cats = tree_meta[grp]
        grp_esc = grp.replace('"', '&quot;').replace("'", "&#39;")
        cat_items = []
        for ci, cat in enumerate(cats):
            cat_esc = cat.replace('"', '&quot;').replace("'", "&#39;")
            cat_items.append(
                f'<div class="tree-cat-item" data-gi="{gi}" data-ci="{ci}" '
                f'data-grp="{grp_esc}" data-cat="{cat_esc}">'
                f'<input type="checkbox" class="tree-cat-cb" id="g{gi}c{ci}" '
                f'data-gi="{gi}" data-grp="{grp_esc}" data-cat="{cat_esc}" '
                f'onchange="onCatChange({gi},{ci})" checked>'
                f'<label for="g{gi}c{ci}">{cat}</label>'
                f'<button type="button" class="tree-sku-btn" '
                f'onclick="openSkuModal({gi},{ci},&quot;{cat_esc}&quot;)" '
                f'title="SKU 선택">▸</button></div>')
        # tree-cats에 closed 추가(기본 접힘), 화살표 span에만 onclick
        sb_parts.append(
            f'<div class="tree-group" id="tg{gi}" data-grp="{grp_esc}">'
            f'<div class="tree-grp-hdr">'
            f'<input type="checkbox" class="tree-grp-cb" id="g{gi}" '
            f'data-gi="{gi}" onchange="onGrpChange({gi})" checked>'
            f'<span class="tree-grp-lbl">{grp}</span>'
            f'<span class="tree-arr closed" id="ta{gi}" onclick="toggleGrp({gi})" title="펼치기/접기">▾</span></div>'
            f'<div class="tree-cats closed" id="tc{gi}">{"".join(cat_items)}</div></div>')
    sidebar_html = ''.join(sb_parts)

    html = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=1280,initial-scale=1.0,shrink-to-fit=yes">
<title>품목별 매출 대시보드 | CMS Lab</title>
<style>{CSS_ITEM}</style>
{chartjs_tag}
</head>
<body>

<div class="sticky-header">
<div class="header" style="background:#fff;padding:0 clamp(12px,2vw,24px)">
  <div class="hdr-company" style="color:#833C0C;cursor:pointer" onclick="resetFilters()" title="초기화">씨엠에스랩</div>
  <div class="hdr-divider" style="background:#e5e7eb"></div>
  <div class="hdr-tabs">
    <button class="hdr-tab active" id="tab-item">📦 품목별 매출</button>
  </div>
  <div class="hdr-right"><strong>{base_date} 기준</strong></div>
</div>
</div>

<div class="sticky-filterbar">
<div class="filterbar">
  <div class="filter-row">
    <div class="fg team-fg">
      <span class="fg-label toggleable" id="teamLabel" onclick="toggleSingleMode('team')" title="클릭 시 단일 선택 모드 전환">팀</span>
      <button class="pb active" data-f="team" data-v="ALL"        onclick="togTeam(this)">전사</button>
      <span class="fg-sep">ㅣ</span>
      <button class="pb" data-f="team" data-v="RBD1팀"           onclick="togTeam(this)">RBD1팀</button>
      <button class="pb" data-f="team" data-v="RBD2팀"           onclick="togTeam(this)">RBD2팀</button>
      <button class="pb" data-f="team" data-v="일본사업팀"        onclick="togTeam(this)">일본사업팀</button>
      <button class="pb" data-f="team" data-v="중국사업팀"        onclick="togTeam(this)">중국사업팀</button>
      <button class="pb" data-f="team" data-v="동북아MC팀"       onclick="togTeam(this)">동북아MC팀</button>
      <span class="fg-sep bu-sep">ㅣ</span>
      <button class="pb" data-f="team" data-v="Global사업팀"      onclick="togTeam(this)">글로벌사업팀</button>
      <button class="pb" data-f="team" data-v="GEC팀"            onclick="togTeam(this)">GEC팀</button>
      <span class="fg-sep bu-sep">ㅣ</span>
      <button class="pb" data-f="team" data-v="메디컬팀"          onclick="togTeam(this)">메디컬팀</button>
    </div>
    <div class="fg">
      <span class="fg-label" id="yearGroup">연도</span>
      <button class="pb" data-f="year" data-v="2024" onclick="togYear(this)">24년 실적</button>
      <button class="pb active" data-f="year" data-v="2025" onclick="togYear(this)">25년 실적</button>
      <button class="pb active" data-f="year" data-v="2026" onclick="togYear(this)">26년 실적</button>
    </div>
    <div class="fg">
      <span class="fg-label" id="brandGroup">브랜드</span>
      <div class="ms-drop" id="brDrop" data-key="brands">
        <button type="button" class="ms-btn" id="brBtn" onclick="toggleMsDrop('brDrop')"><span class="ms-label">전체</span><span class="ms-arrow">▼</span></button>
        <div class="ms-menu">
          <label class="ms-item ms-item-all"><input type="checkbox" value="ALL" checked onchange="onMsChange(this)"><span>전체</span></label>
          {brand_items}
        </div>
      </div>
    </div>
    <div class="fg">
      <span class="fg-label">채널</span>
      <div class="ms-drop" id="chDrop" data-key="channels">
        <button type="button" class="ms-btn" id="chBtn" onclick="toggleMsDrop('chDrop')"><span class="ms-label">전체</span><span class="ms-arrow">▼</span></button>
        <div class="ms-menu" id="chMenu">
          <label class="ms-item ms-item-all"><input type="checkbox" value="ALL" checked onchange="onMsChange(this)"><span>전체</span></label>
          <!-- 채널 항목은 JS에서 팀 선택에 따라 동적 갱신 -->
        </div>
      </div>
    </div>
    <div class="fg" id="countryFg">
      <span class="fg-label">국가</span>
      <div class="ms-drop" id="ctDrop" data-key="countries">
        <button type="button" class="ms-btn" id="ctBtn" onclick="toggleMsDrop('ctDrop')"><span class="ms-label">전체</span><span class="ms-arrow">▼</span></button>
        <div class="ms-menu">
          <label class="ms-item ms-item-all"><input type="checkbox" value="ALL" checked onchange="onMsChange(this)"><span>전체</span></label>
          {country_items}
        </div>
      </div>
    </div>
    <div class="fg">
      <span class="fg-label">거래처</span>
      <div class="ms-drop" id="csDrop" data-key="customers">
        <button type="button" class="ms-btn" id="csBtn" onclick="toggleMsDrop('csDrop')"><span class="ms-label">전체</span><span class="ms-arrow">▼</span></button>
        <div class="ms-menu">
          <label class="ms-item ms-item-all"><input type="checkbox" value="ALL" checked onchange="onMsChange(this)"><span>전체</span></label>
          {customer_items}
        </div>
      </div>
    </div>
  </div>
  <div class="filter-row" style="border-top:1px solid rgba(255,255,255,.15);padding-top:5px;margin-top:2px">
    <div class="fg">
      <span class="fg-label toggleable" id="monthLabel" onclick="toggleSingleMode('month')" title="클릭 시 단일 선택 모드 전환">월</span>
      <button class="pb active" data-f="month" data-v="YTD" onclick="togMonth(this)">YTD</button>
      <span class="fg-sep">ㅣ</span>
      <button class="pb" data-f="month" data-v="1"  onclick="togMonth(this)">1월</button>
      <button class="pb" data-f="month" data-v="2"  onclick="togMonth(this)">2월</button>
      <button class="pb" data-f="month" data-v="3"  onclick="togMonth(this)">3월</button>
      <button class="pb" data-f="month" data-v="4"  onclick="togMonth(this)">4월</button>
      <button class="pb" data-f="month" data-v="5"  onclick="togMonth(this)">5월</button>
      <button class="pb" data-f="month" data-v="6"  onclick="togMonth(this)">6월</button>
      <button class="pb" data-f="month" data-v="7"  onclick="togMonth(this)">7월</button>
      <button class="pb" data-f="month" data-v="8"  onclick="togMonth(this)">8월</button>
      <button class="pb" data-f="month" data-v="9"  onclick="togMonth(this)">9월</button>
      <button class="pb" data-f="month" data-v="10" onclick="togMonth(this)">10월</button>
      <button class="pb" data-f="month" data-v="11" onclick="togMonth(this)">11월</button>
      <button class="pb" data-f="month" data-v="12" onclick="togMonth(this)">12월</button>
    </div>
    <div class="fg">
      <span class="fg-label toggleable" id="quarterGroup" onclick="toggleSingleMode('quarter')" title="클릭 시 단일 선택 모드 전환">분기</span>
      <button class="pb" data-f="quarter" data-v="Q1" onclick="togQuarter(this)">1Q</button>
      <button class="pb" data-f="quarter" data-v="Q2" onclick="togQuarter(this)">2Q</button>
      <button class="pb" data-f="quarter" data-v="Q3" onclick="togQuarter(this)">3Q</button>
      <button class="pb" data-f="quarter" data-v="Q4" onclick="togQuarter(this)">4Q</button>
    </div>
    <div class="fg" id="saleTypeFg" data-tab-only="item">
      <span class="fg-label" id="saleTypeLabel" style="background:#FFFACD;color:#7c5800;border-color:#e6c800">판매구분</span>
      <button class="pb active" data-f="st" data-v="ALL"  onclick="togST(this)">전체</button>
      <button class="pb" data-f="st" data-v="판매품"       onclick="togST(this)">판매품</button>
      <button class="pb" data-f="st" data-v="증정품"       onclick="togST(this)">증정품</button>
    </div>
    <div class="fg" id="themeFg">
      <span class="fg-label">테마</span>
      <div class="ms-drop" id="thDrop" data-key="themes">
        <button type="button" class="ms-btn" id="thBtn" onclick="toggleMsDrop('thDrop')"><span class="ms-label">전체</span><span class="ms-arrow">▼</span></button>
        <div class="ms-menu">
          <label class="ms-item ms-item-all"><input type="checkbox" value="ALL" checked onchange="onMsChange(this)"><span>전체</span></label>
          {theme_items}
        </div>
      </div>
    </div>
  </div>
</div>
</div>

<!-- ═════ 품목별 매출 탭 (기본 노출) ═════ -->
<div class="tab-pane" id="pane-item">
<div class="item-kpi-strip" id="itemKpiStrip">
  <!-- 그룹 1: 매출액 (증감 텍스트 짧아 3등분 유지) -->
  <div class="kpi-group">
    <div class="kpi-group-hdr">매출액 <span>(백만원)</span></div>
    <div class="kpi-group-body">
      <div class="kpi-card"><div class="kpi-t">2025년</div><div class="kpi-v" id="kRev25">-</div></div>
      <div class="kpi-card"><div class="kpi-t">2026년</div><div class="kpi-v" id="kRev26">-</div></div>
      <div class="kpi-card kpi-diff neu" id="kRevDiffBox"><div class="kpi-t">증감</div><div class="kpi-v" id="kRevDiff">-</div></div>
    </div>
  </div>

  <!-- 매출비중 KPI (매출액 그룹 오른쪽) — 2026년 필터된 매출 / 2026년 기간 총 매출 -->
  <div class="kpi-group kpi-single">
    <div class="kpi-group-hdr">2026년 매출비중 <span>(%)</span></div>
    <div class="kpi-group-body">
      <div class="kpi-card"><div class="kpi-t">선택 조건 매출</div><div class="kpi-v" id="kRevShare">-</div></div>
    </div>
  </div>

  <!-- 그룹 2: 매출총이익률 (증감 %p로 짧아 3등분 유지) -->
  <div class="kpi-group">
    <div class="kpi-group-hdr">매출총이익률</div>
    <div class="kpi-group-body">
      <div class="kpi-card"><div class="kpi-t">2025년</div><div class="kpi-v" id="kCR25">-</div></div>
      <div class="kpi-card"><div class="kpi-t">2026년</div><div class="kpi-v" id="kCR26">-</div></div>
      <div class="kpi-card kpi-diff neu" id="kCRDiffBox"><div class="kpi-t">증감</div><div class="kpi-v" id="kCRDiff">-</div></div>
    </div>
  </div>

  <!-- 그룹 3: 판매수량 -->
  <div class="kpi-group has-diff">
    <div class="kpi-group-hdr">판매수량 <span>(개)</span></div>
    <div class="kpi-group-body">
      <div class="kpi-card"><div class="kpi-t">2025년</div><div class="kpi-v" id="kQty25">-</div></div>
      <div class="kpi-card"><div class="kpi-t">2026년</div><div class="kpi-v" id="kQty26">-</div></div>
      <div class="kpi-card kpi-diff neu" id="kQtyDiffBox"><div class="kpi-t">증감</div><div class="kpi-v" id="kQtyDiff">-</div></div>
    </div>
  </div>

  <!-- 그룹 4: 품목정보 (필터 적용 후 유동) -->
  <div class="kpi-group">
    <div class="kpi-group-hdr">품목정보</div>
    <div class="kpi-group-body">
      <div class="kpi-card"><div class="kpi-t">품목군 수</div><div class="kpi-v" id="kGrpCnt">-</div></div>
      <div class="kpi-card"><div class="kpi-t">품목분류 수</div><div class="kpi-v" id="kCatCnt">-</div></div>
      <div class="kpi-card"><div class="kpi-t">SKU 개수</div><div class="kpi-v" id="kSkuCnt">-</div></div>
    </div>
  </div>

</div>

<!-- SKU 선택 모달 -->
<div class="sku-modal-bg" id="skuModalBg" onclick="closeSkuModal()"></div>
<div class="sku-modal" id="skuModal">
  <div class="sku-modal-hdr">
    <span id="skuModalTitle">품목분류</span>
    <button type="button" class="sku-modal-close" onclick="closeSkuModal()" title="닫기">✕</button>
  </div>
  <div class="sku-modal-toolbar">
    <button type="button" class="sb-btn" onclick="setSkuAll(true)">전체</button>
    <button type="button" class="sb-btn" onclick="setSkuAll(false)">해제</button>
    <span class="sku-modal-count" id="skuModalCount"></span>
  </div>
  <div class="sku-modal-body" id="skuModalBody"></div>
</div>

<div class="layout">
  <div class="sidebar" id="sidebar">
    <div class="sidebar-hdr sidebar-hdr-2rows">
      <span class="sb-title">품목군 / 품목분류</span>
      <div class="sb-btns-col">
        <div class="sb-btns-row">
          <button class="sb-btn sb-btn-a" onclick="setSbAll(true)"  title="모두 선택">전체</button>
          <button class="sb-btn sb-btn-b" onclick="setSbAll(false)" title="모두 해제">해제</button>
        </div>
        <div class="sb-btns-row">
          <button class="sb-btn sb-btn-a" onclick="expandAllTree()"   title="모든 품목군 펼치기">▾ 펼치기</button>
          <button class="sb-btn sb-btn-b" onclick="collapseAllTree()" title="모두 접기">▸ 닫기</button>
        </div>
      </div>
    </div>
    {sidebar_html}
  </div>
  <div class="content">
    <div class="charts-full">
      <div class="chart-card" id="chartRevCard">
        <div class="chart-card-hdr">
          <span class="chart-card-title title-rev">월별 매출액 (백만원)</span>
          <div class="chart-legend" id="legendRev"></div>
          <button class="chart-expand-btn" id="chartRevExpandBtn" onclick="toggleChartExpand('chartRevCard','chartRevExpandBtn')" title="확대">⛶</button>
        </div>
        <div class="chart-card-body" style="height:300px"><canvas id="chartRev"></canvas></div>
        <button class="stbl-toggle" onclick="toggleStbl('stblRev',this)">
          <span>월별 상세 데이터 (백만원)</span><span class="stbl-arr">▾</span>
        </button>
        <div class="stbl-wrap" id="stblRev">
          <div style="overflow-x:auto;padding:0 14px 8px"><table class="stbl" id="stblRevTbl"></table></div>
        </div>
      </div>
      <!-- 누적 차트 2개 (반폭) -->
      <div class="charts-half">
        <div class="chart-card" id="chartCumRevCard">
          <div class="chart-card-hdr">
            <span class="chart-card-title title-rev">누적 매출액 (백만원)</span>
            <div class="chart-legend" id="legendCumRev"></div>
            <button class="chart-expand-btn" id="chartCumRevExpandBtn" onclick="toggleChartExpand('chartCumRevCard','chartCumRevExpandBtn')" title="확대">⛶</button>
          </div>
          <div class="chart-card-body" style="height:300px"><canvas id="chartCumRev"></canvas></div>
          <button class="stbl-toggle" onclick="toggleStbl('stblCumRev',this)">
            <span>월별 상세 데이터 (백만원)</span><span class="stbl-arr">▾</span>
          </button>
          <div class="stbl-wrap" id="stblCumRev">
            <div style="overflow-x:auto;padding:0 14px 8px"><table class="stbl" id="stblCumRevTbl"></table></div>
          </div>
        </div>
        <div class="chart-card" id="chartCumQtyCard">
          <div class="chart-card-hdr">
            <span class="chart-card-title title-qty">누적 매출수량 (개)</span>
            <div class="chart-legend" id="legendCumQty"></div>
            <button class="chart-expand-btn" id="chartCumQtyExpandBtn" onclick="toggleChartExpand('chartCumQtyCard','chartCumQtyExpandBtn')" title="확대">⛶</button>
          </div>
          <div class="chart-card-body" style="height:300px"><canvas id="chartCumQty"></canvas></div>
          <button class="stbl-toggle" onclick="toggleStbl('stblCumQty',this)">
            <span>월별 상세 데이터 (개)</span><span class="stbl-arr">▾</span>
          </button>
          <div class="stbl-wrap" id="stblCumQty">
            <div style="overflow-x:auto;padding:0 14px 8px"><table class="stbl" id="stblCumQtyTbl"></table></div>
          </div>
        </div>
      </div>
      <div class="chart-card" id="chartQtyCard">
        <div class="chart-card-hdr">
          <span class="chart-card-title title-qty">월별 매출수량 (개)</span>
          <div class="chart-legend" id="legendQty"></div>
          <button class="chart-expand-btn" id="chartQtyExpandBtn" onclick="toggleChartExpand('chartQtyCard','chartQtyExpandBtn')" title="확대">⛶</button>
        </div>
        <div class="chart-card-body" style="height:300px"><canvas id="chartQty"></canvas></div>
        <button class="stbl-toggle" onclick="toggleStbl('stblQty',this)">
          <span>월별 상세 데이터 (개)</span><span class="stbl-arr">▾</span>
        </button>
        <div class="stbl-wrap" id="stblQty">
          <div style="overflow-x:auto;padding:0 14px 8px"><table class="stbl" id="stblQtyTbl"></table></div>
        </div>
      </div>
      <!-- Top5 채널 월별 매출 추이 (월별 매출수량 아래로 이동, 표 위) -->
      <div class="chart-card" id="chartTop5Card">
        <div class="chart-card-hdr">
          <span class="chart-card-title" style="color:#833C0C">Top5 채널 월별 매출 추이 (백만원)</span>
          <div class="chart-legend" id="legendTop5"></div>
          <button class="chart-expand-btn" id="chartTop5ExpandBtn" onclick="toggleChartExpand('chartTop5Card','chartTop5ExpandBtn')" title="확대">⛶</button>
        </div>
        <div class="chart-card-body" style="height:300px"><canvas id="chartTop5"></canvas></div>
        <button class="stbl-toggle" onclick="toggleStbl('stblTop5',this)">
          <span>월별 상세 데이터 (백만원)</span><span class="stbl-arr">▾</span>
        </button>
        <div class="stbl-wrap" id="stblTop5">
          <div style="overflow-x:auto;padding:0 14px 8px"><table class="stbl" id="stblTop5Tbl"></table></div>
        </div>
      </div>
    </div>
    <div class="tbl-card">
      <div class="tbl-toolbar">
        <span class="tbl-title">품목별 매출 현황</span>
        <span class="tbl-count" id="tblCount"></span>
        <button class="sb-btn tbl-expand-btn" onclick="expandAllTbl1()" title="모든 품목군 → 품목분류까지 펼치기">▾ 펼치기</button>
        <button class="sb-btn tbl-expand-btn" onclick="collapseAllTbl1()" title="모두 접기">▸ 닫기</button>
        <input type="text" class="tbl-search" id="tblSearch" placeholder="품목군/분류/SKU 검색..." oninput="renderTable(lastRows)">
      </div>
      <div class="tbl-outer" id="tblOuter"></div>
    </div>
    <!-- 두 번째 표: 채널 > 거래처 > 품목군 > 품목분류 > SKU 계층 -->
    <div class="tbl-card" style="margin-top:12px">
      <div class="tbl-toolbar tbl-toolbar-2">
        <span class="tbl-title">채널별 매출 현황</span>
        <span class="tbl-count" id="tblCount2"></span>
        <button class="sb-btn tbl-expand-btn" onclick="expandAllTbl2()" title="채널→거래처→품목군→품목분류까지 펼치기">▾ 펼치기</button>
        <button class="sb-btn tbl-expand-btn" onclick="collapseAllTbl2()" title="모두 접기">▸ 닫기</button>
        <input type="text" class="tbl-search" id="tblSearch2" placeholder="채널/거래처/품목 검색..." oninput="renderTable2(lastRows)">
      </div>
      <div class="tbl-outer" id="tblOuter2"></div>
    </div>
  </div>
</div>
</div>
<!-- ═════ /pane-item ═════ -->

<script>
'use strict';
{raw_decl}
const TREE = {tree_json};
const GRP_ORDER = {grp_order_json};
const SKU_MAP = {sku_map_json};   // (gi_ci) → [{{sku, sku_name}}]
const CHANNEL_ORDER = {channel_order_json};   // 팀참고 파일 채널 순서
const BRAND_ORDER   = ['CFC','CEX','DMB','SUS','PIN','KTZ','ETC'];  // 브랜드 고정 순서

const YEAR_ORDER  = ['2024','2025','2026'];
const YEAR_COLORS = {{'2024':'#6b7280','2025':'#5a96c8','2026':'#5C2508'}};
const YEAR_LABELS = {{'2024':'2024년','2025':'2025년','2026':'2026년'}};

const S = {{
  teams:     new Set(['ALL']),
  years:     new Set(['2025','2026']),
  months:    new Set(['YTD']),
  channels:  new Set(['ALL']),
  brands:    new Set(['ALL']),
  countries: new Set(['ALL']),
  customers: new Set(['ALL']),
  themes:    new Set(['ALL']),
  saleType:  'ALL',
  singleTeam:    false,
  singleMonth:   false,
  singleQuarter: false,
}};
let lastRows=[],revChart=null,qtyChart=null,cumRevChart=null,cumQtyChart=null,top5Chart=null;
const uncheckedSkus = new Set();   // 사용자가 SKU 모달에서 해제한 SKU 코드
let currentSkuModal = null;         // {{gi, ci}} — 현재 열려있는 모달의 품목분류

function activeYears(){{ return YEAR_ORDER.filter(y=>S.years.has(y)); }}

/* ─── 카스케이딩 다중선택 드롭다운 ─────
   각 드롭다운(브랜드/채널/국가/거래처)의 옵션은 "자기 자신을 제외한 다른 필터들"을 통과하는
   레코드의 고유값으로 좁혀짐. → 상호 종속적으로 옵션이 축소됨 */
function computeAvailableOptions(){{
  const brands=new Set(), channels=new Set(), countries=new Set(), customers=new Set(), themes=new Set();
  RAW.forEach(d => {{
    const teamOk    = S.teams.has('ALL')     || S.teams.has(d.team);
    if(!teamOk) return;
    const brandOk   = S.brands.has('ALL')    || S.brands.has(d.brand);
    const chanOk    = S.channels.has('ALL')  || S.channels.has(d.channel);
    const ctryOk    = S.countries.has('ALL') || S.countries.has(d.country);
    const custOk    = S.customers.has('ALL') || S.customers.has(d.customer);
    const themeOk   = S.themes.has('ALL')    || S.themes.has(d.theme);
    // "자기 자신을 제외한 나머지 필터"가 통과할 때만 옵션 노출 (카스케이딩)
    if(chanOk && ctryOk && custOk && themeOk && d.brand)    brands.add(d.brand);
    if(brandOk && ctryOk && custOk && themeOk && d.channel) channels.add(d.channel);
    if(brandOk && chanOk && custOk && themeOk && d.country) countries.add(d.country);
    if(brandOk && chanOk && ctryOk && themeOk && d.customer)customers.add(d.customer);
    if(brandOk && chanOk && ctryOk && custOk && d.theme)    themes.add(d.theme);
  }});
  return {{brands, channels, countries, customers, themes}};
}}

/* 주어진 order 배열 순서 우선, 나머지는 뒤에 한글순 */
function sortByOrder(arr, order){{
  const idx = new Map();
  order.forEach((v,i)=>idx.set(v,i));
  return arr.slice().sort((a,b)=>{{
    const ai = idx.has(a) ? idx.get(a) : 1e9;
    const bi = idx.has(b) ? idx.get(b) : 1e9;
    if(ai !== bi) return ai - bi;
    return a.localeCompare(b,'ko');
  }});
}}
function sortByChannelOrder(arr){{ return sortByOrder(arr, CHANNEL_ORDER); }}
function sortByBrandOrder(arr)  {{ return sortByOrder(arr, BRAND_ORDER); }}

/* 특정 드롭다운의 옵션 재구성 (기존 체크 상태 보존) */
function refreshMsDropdown(dropId, availableSet, setKey){{
  const menu = document.querySelector('#' + dropId + ' .ms-menu');
  if(!menu) return;
  const S_set = S[setKey];
  const sorted = setKey === 'channels' ? sortByChannelOrder([...availableSet])
               : setKey === 'brands'   ? sortByBrandOrder([...availableSet])
               : [...availableSet].sort((a,b)=>a.localeCompare(b,'ko'));
  let html = '<label class="ms-item ms-item-all"><input type="checkbox" value="ALL"'
           + (S_set.has('ALL')?' checked':'')
           + ' onchange="onMsChange(this)"><span>전체</span></label>';
  sorted.forEach(v => {{
    const ck = S_set.has(v) && !S_set.has('ALL');
    const safeVal = String(v).replace(/"/g,'&quot;');
    html += '<label class="ms-item"><input type="checkbox" value="'+safeVal+'"'
         +  (ck?' checked':'')+' onchange="onMsChange(this)"><span>'+v+'</span></label>';
  }});
  menu.innerHTML = html;
  // 옵션에서 사라진 값은 S에서 제거
  [...S_set].forEach(v => {{ if(v !== 'ALL' && !availableSet.has(v)) S_set.delete(v); }});
  if(S_set.size === 0) S_set.add('ALL');
  updateMsBtnLabel(setKey);
}}

/* 현재 열려있는(방금 사용자가 조작한) 드롭다운을 제외한 나머지 드롭다운을 새로 계산 */
function refreshOtherDropdowns(exceptId){{
  const opts = computeAvailableOptions();
  const config = [
    ['brDrop', opts.brands,    'brands'],
    ['chDrop', opts.channels,  'channels'],
    ['ctDrop', opts.countries, 'countries'],
    ['csDrop', opts.customers, 'customers'],
    ['thDrop', opts.themes,    'themes'],
  ];
  config.forEach(([id, set, key]) => {{
    if(id !== exceptId) refreshMsDropdown(id, set, key);
  }});
}}

/* 팀 변경 등 외부 트리거 시 모든 드롭다운 새로 계산 */
function refreshAllDropdowns(){{
  refreshOtherDropdowns(null);
}}

/* (removed) updateChannelDropdown wrapper — 직접 refreshAllDropdowns 호출로 대체 */

/* ─── 다중선택 드롭다운 공통 함수 ─── */
function toggleMsDrop(id){{
  const el = document.getElementById(id);
  if (!el) return;
  // 다른 드롭다운은 닫기
  document.querySelectorAll('.ms-drop.open').forEach(d => {{ if (d.id !== id) d.classList.remove('open'); }});
  el.classList.toggle('open');
}}

function onMsChange(cb){{
  const drop = cb.closest('.ms-drop');
  if (!drop) return;
  const key  = drop.dataset.key;     // brands / channels / countries
  const set  = S[key];
  const val  = cb.value;
  const menu = drop.querySelector('.ms-menu');
  const allCb= menu.querySelector('input[value="ALL"]');

  if (val === 'ALL'){{
    // 전체 항목: 언체크 방지 (최소 하나는 선택)
    if (!cb.checked){{ cb.checked = true; return; }}
    // 전체 체크: S를 ALL만 남기고 다른 체크 해제
    set.clear(); set.add('ALL');
    menu.querySelectorAll('input[type=checkbox]').forEach(c => {{ if (c.value !== 'ALL') c.checked = false; }});
  }} else {{
    if (cb.checked){{
      set.delete('ALL');
      if (allCb) allCb.checked = false;
      set.add(val);
    }} else {{
      set.delete(val);
      // 개별 항목이 모두 해제되면 전체로 복귀
      if (![...set].some(v => v !== 'ALL')){{
        set.clear(); set.add('ALL');
        if (allCb) allCb.checked = true;
      }}
    }}
  }}
  updateMsBtnLabel(key);
  // 다른 드롭다운의 옵션을 카스케이딩으로 좁힘 (현재 조작한 드롭다운은 유지)
  refreshOtherDropdowns(drop.id);
  render();
}}

function updateMsBtnLabel(key){{
  const idMap = {{brands:'brBtn', channels:'chBtn', countries:'ctBtn', customers:'csBtn', themes:'thBtn'}};
  const btn = document.getElementById(idMap[key]);
  if (!btn) return;
  const set = S[key];
  const lbl = btn.querySelector('.ms-label');
  if (!lbl) return;
  if (set.has('ALL') || set.size === 0)      lbl.textContent = '전체';
  else if (set.size === 1)                    lbl.textContent = [...set][0];
  else                                         lbl.textContent = set.size + '개 선택';
}}

/* 드롭다운 바깥 클릭 시 닫기 */
document.addEventListener('click', (e) => {{
  if (!e.target.closest('.ms-drop')){{
    document.querySelectorAll('.ms-drop.open').forEach(d => d.classList.remove('open'));
  }}
}});

/* ── 숫자 포맷 ── */
/* 음수는 회계 표기법 관례에 따라 괄호로: -3,500 → (3,500) */
function _wrapNeg(strAbs, neg){{ return neg ? '('+strAbs+')' : strAbs; }}

function fmtRev(n){{
  // 매출액: 백만원 단위. 음수는 괄호 표기.
  if(!n) return '-';
  const v = Math.round(n/1e6);
  if(v === 0) return '-';
  return _wrapNeg(Math.abs(v).toLocaleString(), v < 0);
}}
function fmtQ(n){{
  if(!n) return '-';
  const v = Math.round(n);
  if(v === 0) return '-';
  return _wrapNeg(Math.abs(v).toLocaleString(), v < 0);
}}
function fmtCR(rev,cost){{
  // 매출총이익률 = (rev-cost)/rev, 소수점 첫째자리. 음수는 괄호 표기.
  if(!rev) return '-';
  const v = (rev-cost)/rev*100;
  const abs = Math.abs(v).toFixed(1);
  return _wrapNeg(abs+'%', v < 0);
}}

/* ── 필터 ── */
function basePass(d){{
  if(!S.teams.has('ALL')      && !S.teams.has(d.team))          return false;
  if(!S.years.has(String(d.yr)))                                return false;
  if(!S.brands.has('ALL')     && !S.brands.has(d.brand))        return false;
  if(!S.channels.has('ALL')   && !S.channels.has(d.channel))    return false;
  if(!S.countries.has('ALL')  && !S.countries.has(d.country))   return false;
  if(!S.customers.has('ALL')  && !S.customers.has(d.customer))  return false;
  if(!S.themes.has('ALL')     && !S.themes.has(d.theme))        return false;
  if(S.saleType!=='ALL' && d.sale_type!==S.saleType)            return false;
  if(uncheckedSkus.has(d.sku))                                  return false;
  return true;
}}
function getSbUnchecked(){{
  const s=new Set();
  document.querySelectorAll('.tree-cat-cb:not(:checked)').forEach(cb=>s.add(cb.dataset.cat));
  return s;
}}
function fullPass(d,unchecked,addMQ){{
  if(!basePass(d)) return false;
  if(unchecked.size && unchecked.has(d.item_cat)) return false;
  if(addMQ){{
    if(!S.months.has('YTD') && !S.months.has(d.month)) return false;
  }}
  return true;
}}

/* ── 팀 토글 (채널 드롭다운 연동) ── */
function togTeam(btn){{
  const v=btn.dataset.v;
  if(v==='ALL'){{
    S.teams=new Set(['ALL']);
  }} else if(S.singleTeam){{
    if(S.teams.size===1 && S.teams.has(v)){{
      S.teams=new Set(['ALL']);
    }} else {{
      S.teams=new Set([v]);
    }}
  }} else {{
    S.teams.delete('ALL');
    S.teams.has(v) ? S.teams.delete(v) : S.teams.add(v);
    if(!S.teams.size) S.teams=new Set(['ALL']);
  }}
  document.querySelectorAll('[data-f="team"]').forEach(b=>
    b.classList.toggle('active', b.dataset.v==='ALL' ? S.teams.has('ALL') : S.teams.has(b.dataset.v)));
  // 팀 변경 시 브랜드/채널/국가/거래처 선택 상태 유지하고 옵션만 재계산 (카스케이딩)
  refreshAllDropdowns();
  render();
}}

/* ── 채널 드롭다운 (팀 자동 연동) ── */
/* togChannel/togBrand: 복수선택 드롭다운으로 대체됨 (onMsChange 사용) */

/* ── 연도 토글 ── */
function togYear(btn){{
  const v=btn.dataset.v;
  S.years.has(v) ? S.years.delete(v) : S.years.add(v);
  if(!S.years.size){{ S.years.add(v); }}
  document.querySelectorAll('[data-f="year"]').forEach(b=>
    b.classList.toggle('active', S.years.has(b.dataset.v)));
  render();
}}

/* ── 월/분기 (단일 Set 공유) ── */
const Q_MONTHS = {{ Q1:[1,2,3], Q2:[4,5,6], Q3:[7,8,9], Q4:[10,11,12] }};

function syncMonthButtons(){{
  document.querySelectorAll('[data-f="month"]').forEach(b=>{{
    const v=b.dataset.v;
    b.classList.toggle('active', v==='YTD' ? S.months.has('YTD') : S.months.has(Number(v)));
  }});
  document.querySelectorAll('[data-f="quarter"]').forEach(b=>{{
    const qm=Q_MONTHS[b.dataset.v];
    b.classList.toggle('active', qm.every(m=>S.months.has(m)));
  }});
}}

function togMonth(btn){{
  const v=btn.dataset.v;
  if(v==='YTD'){{
    S.months=new Set(['YTD']);
  }} else {{
    const n=Number(v);
    if(S.singleMonth){{
      if(S.months.size===1 && S.months.has(n)){{
        S.months=new Set(['YTD']);
      }} else {{
        S.months=new Set([n]);
      }}
    }} else {{
      S.months.delete('YTD');
      S.months.has(n) ? S.months.delete(n) : S.months.add(n);
      if(!S.months.size) S.months=new Set(['YTD']);
    }}
  }}
  syncMonthButtons();
  render();
}}

function togQuarter(btn){{
  const qm=Q_MONTHS[btn.dataset.v];
  if(S.singleQuarter){{
    const exactMatch=S.months.size===qm.length && qm.every(m=>S.months.has(m));
    if(exactMatch){{
      S.months=new Set(['YTD']);
    }} else {{
      S.months=new Set(qm);
    }}
  }} else {{
    const allSelected=qm.every(m=>S.months.has(m));
    S.months.delete('YTD');
    if(allSelected){{
      qm.forEach(m=>S.months.delete(m));
    }} else {{
      qm.forEach(m=>S.months.add(m));
    }}
    if(!S.months.size) S.months=new Set(['YTD']);
  }}
  syncMonthButtons();
  render();
}}

/* ── 판매구분 토글 ── */
function togST(btn){{
  S.saleType=btn.dataset.v;
  document.querySelectorAll('[data-f="st"]').forEach(b=>b.classList.toggle('active',b.dataset.v===S.saleType));
  render();
}}

/* ── 단일선택 모드 ── */
function toggleSingleMode(kind){{
  const key={{team:'singleTeam',month:'singleMonth',quarter:'singleQuarter'}}[kind];
  if(!key) return;
  S[key]=!S[key];
  const labelMap={{team:'teamLabel',month:'monthLabel',quarter:'quarterGroup'}};
  const lbl=document.getElementById(labelMap[kind]);
  if(lbl) lbl.classList.toggle('single-mode',S[key]);
}}

/* ── 사이드바 트리 ── */
function toggleGrp(gi){{
  const tc=document.getElementById('tc'+gi);
  const ta=document.getElementById('ta'+gi);
  const closed=tc.classList.toggle('closed');
  ta.classList.toggle('closed',closed);
}}
function onGrpChange(gi){{
  const grpCb=document.getElementById('g'+gi);
  document.querySelectorAll('.tree-cat-cb[data-gi="'+gi+'"]').forEach(cb=>{{
    cb.checked=grpCb.checked;
    cb.indeterminate=false;
    // 각 카테고리 내부의 SKU 선택 상태도 초기화 (전체 포함 / 전체 제외)
    const parts = cb.id.match(/^g(\\d+)c(\\d+)$/);
    if(parts){{
      const ci=parts[2];
      const key=gi+'_'+ci;
      const skus=SKU_MAP[key]||[];
      skus.forEach(s=>{{
        if(grpCb.checked) uncheckedSkus.delete(s.sku);
        else uncheckedSkus.add(s.sku);
      }});
      const catItem=document.querySelector('.tree-cat-item[data-gi="'+gi+'"][data-ci="'+ci+'"]');
      if(catItem) catItem.classList.remove('partial-sku');
    }}
  }});
  render();
}}
function updateGrpCbState(gi){{
  const cats=[...document.querySelectorAll('.tree-cat-cb[data-gi="'+gi+'"]')];
  const grpCb=document.getElementById('g'+gi);
  const allCk=cats.every(c=>c.checked);
  const noneCk=cats.every(c=>!c.checked);
  const anyPartial=cats.some(c=>c.indeterminate);   // SKU 일부 선택된 카테고리 존재
  grpCb.checked=allCk && !anyPartial;
  grpCb.indeterminate=(!allCk && !noneCk) || (allCk && anyPartial);
}}
function onCatChange(gi, ci){{
  // 사용자가 사이드바에서 품목분류 체크박스 토글 → 해당 카테고리의 SKU 상태 일괄 동기화
  const catCb=document.getElementById('g'+gi+'c'+ci);
  const key=gi+'_'+ci;
  const skus=SKU_MAP[key]||[];
  if(catCb.checked){{
    // 체크 → 이 카테고리의 모든 SKU 포함
    skus.forEach(s=>uncheckedSkus.delete(s.sku));
  }} else {{
    // 해제 → 이 카테고리의 모든 SKU 제외
    skus.forEach(s=>uncheckedSkus.add(s.sku));
  }}
  // partial 표시 초기화 (전체 체크 또는 전체 해제 상태이므로)
  const catItem=document.querySelector('.tree-cat-item[data-gi="'+gi+'"][data-ci="'+ci+'"]');
  if(catItem) catItem.classList.remove('partial-sku');
  updateGrpCbState(gi);
  render();
}}
function setSbAll(v){{
  document.querySelectorAll('.tree-cat-cb,.tree-grp-cb').forEach(cb=>{{cb.checked=v;cb.indeterminate=false;}});
  // 전체/해제 시 SKU 개별 선택 상태도 초기화
  uncheckedSkus.clear();
  document.querySelectorAll('.tree-cat-item.partial-sku').forEach(el=>el.classList.remove('partial-sku'));
  render();
}}

/* 사이드바 모든 품목군 펼치기 (품목분류 노출) */
function expandAllTree(){{
  document.querySelectorAll('.tree-cats.closed').forEach(el => el.classList.remove('closed'));
  document.querySelectorAll('.tree-arr.closed').forEach(el => el.classList.remove('closed'));
}}
/* 사이드바 모든 품목군 접기 */
function collapseAllTree(){{
  document.querySelectorAll('.tree-cats').forEach(el => el.classList.add('closed'));
  document.querySelectorAll('.tree-arr').forEach(el => el.classList.add('closed'));
}}

/* ── SKU 선택 모달 ── */
/* 현재 활성 필터(판매구분/브랜드/채널/국가/팀)를 통과하는 SKU만 골라냄 */
function getFilteredSkusForCat(gi, ci){{
  const key = gi + '_' + ci;
  const allSkus = SKU_MAP[key] || [];
  // RAW를 스캔해서 현재 필터 통과하는 SKU 목록 구성
  const activeSkus = new Set();
  RAW.forEach(d => {{
    if(!S.teams.has('ALL')     && !S.teams.has(d.team))         return;
    if(!S.brands.has('ALL')    && !S.brands.has(d.brand))       return;
    if(!S.channels.has('ALL')  && !S.channels.has(d.channel))   return;
    if(!S.countries.has('ALL') && !S.countries.has(d.country))  return;
    if(!S.customers.has('ALL') && !S.customers.has(d.customer)) return;
    if(!S.themes.has('ALL')    && !S.themes.has(d.theme))       return;
    if(S.saleType !== 'ALL' && d.sale_type !== S.saleType)      return;
    activeSkus.add(d.sku);
  }});
  return allSkus.filter(s => activeSkus.has(s.sku));
}}

function openSkuModal(gi, ci, catName){{
  currentSkuModal = {{gi, ci}};
  const skus = getFilteredSkusForCat(gi, ci);
  document.getElementById('skuModalTitle').textContent = catName + ' — SKU 선택 (' + skus.length + '개)';
  const body = document.getElementById('skuModalBody');
  body.innerHTML = skus.map(s => {{
    const isChecked = !uncheckedSkus.has(s.sku);
    const skuEsc = String(s.sku).replace(/'/g, "\\'");
    // 판매품/증정품 배지 — SKU# 앞에 위치 (긴 이름에 잘리지 않도록)
    let badgeCls = 'sb-none', badgeTxt = s.sale_type || '-';
    if(s.sale_type === '판매품') badgeCls = 'sb-sale';
    else if(s.sale_type === '증정품') badgeCls = 'sb-gift';
    return '<div class="sku-item">'
      + '<input type="checkbox" id="skuCb_' + s.sku + '" data-sku="' + s.sku + '" '
      + (isChecked ? 'checked ' : '') + 'onchange="onSkuChange(\\''+ skuEsc +'\\')">'
      + '<label for="skuCb_' + s.sku + '">'
      + '<span class="sale-badge ' + badgeCls + '">' + badgeTxt + '</span>'
      + '<span class="sku-code">' + s.sku + '</span>'
      + (s.sku_name || '')
      + '</label></div>';
  }}).join('');
  updateSkuModalCount();
  document.getElementById('skuModalBg').classList.add('open');
  document.getElementById('skuModal').classList.add('open');
}}

function closeSkuModal(){{
  document.getElementById('skuModalBg').classList.remove('open');
  document.getElementById('skuModal').classList.remove('open');
  if(currentSkuModal){{
    updateCatSkuState(currentSkuModal.gi, currentSkuModal.ci);
    updateGrpCbState(currentSkuModal.gi);
    currentSkuModal = null;
    render();
  }}
}}

function onSkuChange(sku){{
  const cb = document.querySelector('.sku-item input[data-sku="' + sku + '"]');
  if(!cb) return;
  if(cb.checked) uncheckedSkus.delete(sku);
  else uncheckedSkus.add(sku);
  updateSkuModalCount();
}}

function setSkuAll(v){{
  if(!currentSkuModal) return;
  // 필터를 통과한 SKU(모달에 표시된 SKU)에만 적용
  const skus = getFilteredSkusForCat(currentSkuModal.gi, currentSkuModal.ci);
  skus.forEach(s => {{
    if(v) uncheckedSkus.delete(s.sku);
    else uncheckedSkus.add(s.sku);
  }});
  document.querySelectorAll('.sku-item input').forEach(cb => cb.checked = v);
  updateSkuModalCount();
}}

function updateSkuModalCount(){{
  if(!currentSkuModal) return;
  // 필터된 SKU 기준으로 선택 카운트 표시
  const skus = getFilteredSkusForCat(currentSkuModal.gi, currentSkuModal.ci);
  const selected = skus.filter(s => !uncheckedSkus.has(s.sku)).length;
  document.getElementById('skuModalCount').textContent = '선택 ' + selected + ' / ' + skus.length;
}}

/* 품목분류 아이템의 partial-sku 표시 및 체크박스 indeterminate 상태 갱신 */
function updateCatSkuState(gi, ci){{
  const key = gi + '_' + ci;
  const skus = SKU_MAP[key] || [];
  const totalCount = skus.length;
  const uncheckedCount = skus.filter(s => uncheckedSkus.has(s.sku)).length;
  const catItem = document.querySelector('.tree-cat-item[data-gi="' + gi + '"][data-ci="' + ci + '"]');
  if(!catItem) return;
  const isPartial = uncheckedCount > 0 && uncheckedCount < totalCount;
  const isAllUnchecked = uncheckedCount === totalCount && totalCount > 0;
  catItem.classList.toggle('partial-sku', isPartial);
  const catCb = document.getElementById('g' + gi + 'c' + ci);
  if(catCb){{
    if(isAllUnchecked){{
      catCb.checked = false;
      catCb.indeterminate = false;
    }} else {{
      catCb.checked = true;
      catCb.indeterminate = isPartial;
    }}
  }}
}}

/* ── 초기화 ── */
function resetFilters(){{
  S.teams=new Set(['ALL']);
  S.years=new Set(['2025','2026']);
  S.months=new Set(['YTD']);
  S.channels=new Set(['ALL']);
  S.brands=new Set(['ALL']);
  S.saleType='ALL';
  S.singleTeam=false; S.singleMonth=false; S.singleQuarter=false;
  document.querySelectorAll('[data-f="team"]').forEach(b=>b.classList.toggle('active',b.dataset.v==='ALL'));
  document.querySelectorAll('[data-f="year"]').forEach(b=>b.classList.toggle('active',S.years.has(b.dataset.v)));
  document.querySelectorAll('[data-f="month"]').forEach(b=>b.classList.toggle('active',b.dataset.v==='YTD'));
  document.querySelectorAll('[data-f="quarter"]').forEach(b=>b.classList.remove('active'));
  document.querySelectorAll('[data-f="st"]').forEach(b=>b.classList.toggle('active',b.dataset.v==='ALL'));
  document.getElementById('teamLabel').classList.remove('single-mode');
  document.getElementById('monthLabel').classList.remove('single-mode');
  document.getElementById('quarterGroup').classList.remove('single-mode');
  // 브랜드/채널/국가/거래처/테마 다중선택 드롭다운 초기화 (모두 '전체'로)
  S.brands    = new Set(['ALL']);
  S.channels  = new Set(['ALL']);
  S.countries = new Set(['ALL']);
  S.customers = new Set(['ALL']);
  S.themes    = new Set(['ALL']);
  refreshAllDropdowns();
  updateMsBtnLabel('brands');
  updateMsBtnLabel('channels');
  updateMsBtnLabel('countries');
  updateMsBtnLabel('customers');
  updateMsBtnLabel('themes');
  setSbAll(true);
  alignQuarterToYear();
  updateStickyOffsets();
}}

/* ── 헤더/필터바/KPI 스트립 높이를 실측해 sticky 오프셋(CSS 변수)으로 반영 ── */
function updateStickyOffsets(){{
  const h = document.querySelector('.sticky-header')?.offsetHeight   || 44;
  const f = document.querySelector('.sticky-filterbar')?.offsetHeight || 70;
  const k = document.querySelector('.item-kpi-strip')?.offsetHeight   || 60;
  const t = document.querySelector('.tbl-toolbar')?.offsetHeight      || 42;
  document.documentElement.style.setProperty('--sticky-hf', (h+f)+'px');
  document.documentElement.style.setProperty('--sticky-total', (h+f+k)+'px');
  document.documentElement.style.setProperty('--tbl-toolbar-h', t+'px');
}}

/* ── YTD를 전사에, 분기를 연도에 세로 맞춤 (매출 Dashboard_vf.py와 동일 로직) ── */
function alignQuarterToYear(){{
  requestAnimationFrame(() => requestAnimationFrame(() => {{
    const fb = document.querySelector('.filterbar');
    if (!fb) return;
    const fbLeft = fb.getBoundingClientRect().left;

    const allBtn = document.querySelector('[data-f="team"][data-v="ALL"]');
    const ytdBtn = document.querySelector('[data-f="month"][data-v="YTD"]');
    if (allBtn && ytdBtn) {{
      const sep1 = allBtn.nextElementSibling;
      const sep2 = ytdBtn.nextElementSibling;
      if (sep1 && sep2) {{
        ytdBtn.style.marginLeft = '0px';
        const sep1Left = sep1.getBoundingClientRect().left - fbLeft;
        const sep2Left = sep2.getBoundingClientRect().left - fbLeft;
        ytdBtn.style.marginLeft = Math.max(sep1Left - sep2Left, 0) + 'px';
      }}
    }}

    const yrFg = document.getElementById('yearGroup')?.closest('.fg');
    if (yrFg) yrFg.style.marginLeft = '';

    const yrLabel = document.getElementById('yearGroup');
    const qLabel  = document.getElementById('quarterGroup');
    const lastMon = document.querySelector('[data-f="month"][data-v="12"]');
    if (yrLabel && qLabel && lastMon) {{
      qLabel.style.marginLeft = '0px';
      const yrLeft = yrLabel.getBoundingClientRect().left - fbLeft;
      const qLeft0 = qLabel.getBoundingClientRect().left - fbLeft;
      qLabel.style.marginLeft = Math.max(yrLeft - qLeft0, 12) + 'px';
    }}

    // 판매구분 ← 브랜드 정렬 (판매구분 라벨이 브랜드 라벨과 같은 X 위치에 오도록)
    const brLabel = document.getElementById('brandGroup');
    const stLabel = document.getElementById('saleTypeLabel');
    if (brLabel && stLabel) {{
      stLabel.style.marginLeft = '0px';
      const brLeft = brLabel.getBoundingClientRect().left - fbLeft;
      const stLeft0 = stLabel.getBoundingClientRect().left - fbLeft;
      stLabel.style.marginLeft = Math.max(brLeft - stLeft0, 12) + 'px';
    }}

    // 테마 ← 국가 정렬 (테마가 국가와 같은 X 위치에 오도록, 2행)
    const countryFg = document.getElementById('countryFg');
    const themeFg   = document.getElementById('themeFg');
    if (countryFg && themeFg) {{
      themeFg.style.marginLeft = '0px';
      const cLeft = countryFg.getBoundingClientRect().left - fbLeft;
      const tLeft0 = themeFg.getBoundingClientRect().left - fbLeft;
      themeFg.style.marginLeft = Math.max(cLeft - tLeft0, 12) + 'px';
    }}
  }}));
}}

/* ── 집계 헬퍼 ── */
function sumBy(rows,key){{return rows.reduce((s,d)=>s+d[key],0);}}
function groupBy(rows,key){{
  const m=new Map();
  rows.forEach(d=>{{const k=d[key];if(!m.has(k))m.set(k,[]);m.get(k).push(d);}});
  return m;
}}

/* ── Chart.js 인라인 데이터 레이블 플러그인 ── */
Chart.register({{
  id:'itemDatalabels',
  afterDatasetsDraw(chart){{
    const ctx=chart.ctx;
    const allData=chart.data.datasets.map(ds=>ds.data);
    // 각 x-index 마다 시리즈 값들을 정렬해 상대 순위로 위/아래 배치를 세밀하게 결정
    chart.data.datasets.forEach((ds,di)=>{{
      const meta=chart.getDatasetMeta(di);
      if(meta.hidden) return;
      meta.data.forEach((pt,pi)=>{{
        const v=ds.data[pi];
        if(v===null||v===undefined) return;
        // 해당 x-index의 시리즈 값들 중 유효 값만 수집·정렬
        const otherVals=[];
        for(let oi=0;oi<allData.length;oi++){{
          if(oi===di) continue;
          const ov=allData[oi][pi];
          if(ov!==null&&ov!==undefined) otherVals.push(ov);
        }}
        // 다른 시리즈들 중 자신보다 큰 값 개수 / 작은 값 개수를 세어 밸런스 결정
        let biggerCount=0, smallerCount=0;
        otherVals.forEach(ov=>{{ if(ov>v) biggerCount++; else if(ov<v) smallerCount++; }});
        // 자신이 최소면 아래, 최대면 위. 중간이면 상대적으로 큰 쪽이 위, 작은 쪽이 아래
        //   biggerCount === 0    → 자신이 최대치 → 위
        //   smallerCount === 0   → 자신이 최소치 → 아래
        //   그 외 → 위쪽에 큰 값이 있으면 아래로, 아니면 위로 (자연스러운 분산)
        let above;
        if(biggerCount === 0)       above = true;    // 최고점 → 위
        else if(smallerCount === 0) above = false;   // 최저점 → 아래
        else                        above = (biggerCount <= smallerCount);
        ctx.save();
        ctx.fillStyle=ds.borderColor;
        ctx.font='bold 10px Malgun Gothic,sans-serif';   // 기존 9px → 10px (+1pt)
        ctx.textAlign='center';
        ctx.textBaseline=above?'bottom':'top';
        // 선에서 12px 떨어져서 표시 (기존 8px)
        ctx.fillText(Math.round(v).toLocaleString(), pt.x, pt.y + (above ? -12 : 12));
        ctx.restore();
      }});
    }});
  }}
}});

/* ── KPI 카드 렌더 (15개 개별 카드: 지표 3세트×4 + 카운트 3) ── */
const KPI_YEARS = ['2025','2026'];

function renderKPI(rows){{
  const tot={{}};
  KPI_YEARS.forEach(y=>tot[y]={{rev:0,cost:0,qty:0}});
  // 필터 적용 후 유동 카운트용 (매출/수량이 있는 유효 행만 집계)
  const grpSet=new Set(), catSet=new Set(), skuSet=new Set();
  rows.forEach(d=>{{
    const y=String(d.yr);
    if(tot[y]) {{ tot[y].rev+=d.rev; tot[y].cost+=d.cost; tot[y].qty+=d.qty; }}
    if(d.rev||d.qty){{
      if(d.item_group) grpSet.add(d.item_group);
      if(d.item_cat)   catSet.add(d.item_group+'|'+d.item_cat);
      if(d.sku)        skuSet.add(d.sku);
    }}
  }});

  function setText(id,txt){{
    const el=document.getElementById(id); if(el) el.textContent=txt;
  }}
  // 값 카드 (25/26년) — 단위는 카드 자체에 부착
  //   매출액: 백만원 / 매출총이익률: % (소수점 첫째자리) / 판매수량: 천개 / 품목정보: 개
  setText('kRev25',  tot['2025'].rev ?Math.round(tot['2025'].rev/1e6).toLocaleString():'-');
  setText('kRev26',  tot['2026'].rev ?Math.round(tot['2026'].rev/1e6).toLocaleString():'-');
  setText('kCR25',   tot['2025'].rev ?((tot['2025'].rev-tot['2025'].cost)/tot['2025'].rev*100).toFixed(1)+'%':'-');
  setText('kCR26',   tot['2026'].rev ?((tot['2026'].rev-tot['2026'].cost)/tot['2026'].rev*100).toFixed(1)+'%':'-');
  setText('kQty25',  tot['2025'].qty ?Math.round(tot['2025'].qty).toLocaleString():'-');
  setText('kQty26',  tot['2026'].qty ?Math.round(tot['2026'].qty).toLocaleString():'-');

  // 카운트 카드 (필터 적용 후 유동) — '개' 단위 직접 부착
  setText('kGrpCnt', grpSet.size.toLocaleString()+'개');
  setText('kCatCnt', catSet.size.toLocaleString()+'개');
  setText('kSkuCnt', skuSet.size.toLocaleString()+'개');

  // 증감: 26년 vs 25년 (양수면 증가)
  //   divisor: 표시 단위 배수 (매출액 1e6=백만원, 판매수량 1e3=천개, 나머지 1)
  //   isPct: true면 %p 표기 (매출총이익률)
  const [ya,yb]=KPI_YEARS;
  function setDiff(boxId,valId,getVal,divisor,isPct){{
    const box=document.getElementById(boxId), val=document.getElementById(valId);
    if(!box||!val) return;
    const va=getVal(tot[ya]), vb=getVal(tot[yb]);
    box.className='kpi-card kpi-diff neu';
    if(!va&&!vb){{ val.textContent='-'; return; }}
    const d=vb-va;
    const sign=d>0?'(+)':d<0?'(-)':'';
    const absD=Math.abs(d);
    const absPct=va?Math.abs(d/va*100):null;
    let txt=sign+' ';
    if(isPct){{
      txt+=absD.toFixed(1)+'%p';
    }}else{{
      txt+=Math.round(absD/divisor).toLocaleString();
      if(absPct!==null) txt+=`(${{d>=0?'+':'-'}}${{absPct.toFixed(0)}}%)`;
    }}
    val.textContent=txt.trim();
    box.className='kpi-card kpi-diff '+(d>0?'pos':d<0?'neg':'neu');
  }}
  setDiff('kRevDiffBox','kRevDiff', t=>t.rev,  1e6, false);
  setDiff('kCRDiffBox', 'kCRDiff',  t=>t.rev?((t.rev-t.cost)/t.rev*100):0, 1, true);
  setDiff('kQtyDiffBox','kQtyDiff', t=>t.qty,  1,   false);
}}

/* ── 차트 하단 상세 테이블 토글 ── */
function toggleStbl(id,btn){{
  const wrap=document.getElementById(id);
  const isOpen=wrap.classList.toggle('open');
  btn.classList.toggle('open',isOpen);
}}

/* ── 차트 확대/축소 (매출 Dashboard_vf.py의 확대 로직과 동일) ── */
let _expandedChart=null;   // {{cardId, btnId}} or null
let _expandOverlay=null;

function getExpandRect(){{
  const stickyH = (document.querySelector('.sticky-header')?.offsetHeight   || 44)
                + (document.querySelector('.sticky-filterbar')?.offsetHeight || 70)
                + (document.querySelector('.item-kpi-strip')?.offsetHeight   || 60);
  // 사이드바 우측 이후로 확대 rect를 시작 → 확대 상태에서도 품목군/품목분류 편집 가능
  const sb = document.querySelector('.sidebar');
  const sbRight = sb ? sb.getBoundingClientRect().right : 0;
  const pad = 10;
  const left = Math.max(sbRight + pad, pad);
  return {{ top: stickyH + pad, left: left,
           w: window.innerWidth - left - pad, h: window.innerHeight - stickyH - pad * 2 }};
}}
function applyExpandRect(card){{
  const r=getExpandRect();
  card.style.top=r.top+'px'; card.style.left=r.left+'px';
  card.style.width=r.w+'px'; card.style.height=r.h+'px';
}}
function closeExpandedChart(){{
  if(!_expandedChart) return;
  const card=document.getElementById(_expandedChart.cardId);
  const btn=document.getElementById(_expandedChart.btnId);
  if(card){{ card.classList.remove('chart-expanded'); card.style.top=card.style.left=card.style.width=card.style.height=''; }}
  if(btn){{ btn.textContent='⛶'; btn.title='확대'; }}
  if(_expandOverlay){{ _expandOverlay.remove(); _expandOverlay=null; }}
  _expandedChart=null;
  setTimeout(()=>{{
    if(revChart) revChart.resize();
    if(qtyChart) qtyChart.resize();
    if(cumRevChart) cumRevChart.resize();
    if(cumQtyChart) cumQtyChart.resize();
    if(top5Chart) top5Chart.resize();
  }}, 280);
}}
function toggleChartExpand(cardId,btnId){{
  const card=document.getElementById(cardId);
  const btn=document.getElementById(btnId);
  if(!card||!btn) return;
  if(_expandedChart && _expandedChart.cardId===cardId){{
    closeExpandedChart();
    return;
  }}
  if(_expandedChart) closeExpandedChart();
  card.classList.add('chart-expanded');
  applyExpandRect(card);
  _expandOverlay=document.createElement('div');
  _expandOverlay.className='chart-overlay-bg';
  _expandOverlay.onclick=closeExpandedChart;
  document.body.appendChild(_expandOverlay);
  btn.textContent='✕'; btn.title='축소';
  _expandedChart={{cardId,btnId}};
  setTimeout(()=>{{
    if(revChart) revChart.resize();
    if(qtyChart) qtyChart.resize();
    if(cumRevChart) cumRevChart.resize();
    if(cumQtyChart) cumQtyChart.resize();
    if(top5Chart) top5Chart.resize();
  }}, 280);
}}

/* ── 차트 하단 상세 테이블 렌더 (선택 연도만큼 동적) ── */
const YEAR_CLS = {{'2024':'c24','2025':'c25','2026':'c26'}};

function renderChartTables(baseRows){{
  const ys=activeYears();
  const tmpRev={{}}, tmpQty={{}}, has={{}};
  ys.forEach(y=>{{ tmpRev[y]=new Array(12).fill(0); tmpQty[y]=new Array(12).fill(0); has[y]=new Array(12).fill(false); }});
  baseRows.forEach(d=>{{
    const y=String(d.yr);
    if(!tmpRev[y]) return;
    const mi=d.month-1;
    tmpRev[y][mi]+=d.rev; tmpQty[y][mi]+=d.qty;
    if(d.rev||d.qty) has[y][mi]=true;
  }});
  const ML=['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'];
  function makeTable(tblId,tmp,divisor){{
    const tbl=document.getElementById(tblId);
    if(!tbl) return;
    const hdr='<thead><tr><th>구분</th>'+ML.map(m=>`<th>${{m}}</th>`).join('')+'</tr></thead>';
    let body='<tbody>';
    ys.forEach(y=>{{
      const cls=YEAR_CLS[y];
      body+=`<tr><td class="${{cls}}">${{YEAR_LABELS[y]}}</td>`+
        [...Array(12)].map((_,i)=>{{
          const v=has[y][i]?Math.round(tmp[y][i]/divisor):null;
          return `<td class="${{cls}}">${{v!==null?v.toLocaleString():'-'}}</td>`;
        }}).join('')+'</tr>';
    }});
    body+='</tbody>';
    tbl.innerHTML=hdr+body;
  }}
  makeTable('stblRevTbl',tmpRev,1e6);
  makeTable('stblQtyTbl',tmpQty,1);

  // 누적 상세 데이터: 월별 누적 합계 (데이터 있는 마지막 월 이후는 '-')
  const cumRev={{}}, cumQty={{}}, cumHas={{}};
  ys.forEach(y=>{{
    cumRev[y]=new Array(12).fill(0);
    cumQty[y]=new Array(12).fill(0);
    cumHas[y]=new Array(12).fill(false);
    let lastMon=-1;
    for(let i=0;i<12;i++) if(has[y][i]) lastMon=i;
    let acR=0, acQ=0;
    for(let i=0;i<12;i++){{
      acR += tmpRev[y][i]; acQ += tmpQty[y][i];
      if(i<=lastMon){{ cumRev[y][i]=acR; cumQty[y][i]=acQ; cumHas[y][i]=true; }}
    }}
  }});
  function makeTableCum(tblId,tmp,cumHasArr,divisor){{
    const tbl=document.getElementById(tblId); if(!tbl) return;
    const hdr='<thead><tr><th>구분</th>'+ML.map(m=>`<th>${{m}}</th>`).join('')+'</tr></thead>';
    let body='<tbody>';
    ys.forEach(y=>{{
      const cls=YEAR_CLS[y];
      body+=`<tr><td class="${{cls}}">${{YEAR_LABELS[y]}}</td>`+
        [...Array(12)].map((_,i)=>{{
          const v=cumHasArr[y][i]?Math.round(tmp[y][i]/divisor):null;
          return `<td class="${{cls}}">${{v!==null?v.toLocaleString():'-'}}</td>`;
        }}).join('')+'</tr>';
    }});
    body+='</tbody>';
    tbl.innerHTML=hdr+body;
  }}
  makeTableCum('stblCumRevTbl',cumRev,cumHas,1e6);
  makeTableCum('stblCumQtyTbl',cumQty,cumHas,1);
}}

/* Top5 채널 월별 상세 표 (Top5 채널 × 12개월) — renderCharts에서 호출 */
function renderTop5Table(chNames, chMonthly){{
  const tbl=document.getElementById('stblTop5Tbl'); if(!tbl) return;
  const ML=['1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월'];
  const hdr='<thead><tr><th>채널</th>'+ML.map(m=>`<th>${{m}}</th>`).join('')+'</tr></thead>';
  let body='<tbody>';
  chNames.forEach(ch=>{{
    const arr = chMonthly.get(ch) || new Array(12).fill(0);
    body+=`<tr><td>${{ch}}</td>`+
      arr.map(v=>{{
        const val=v?Math.round(v/1e6):null;
        return `<td>${{val!==null?val.toLocaleString():'-'}}</td>`;
      }}).join('')+'</tr>';
  }});
  body+='</tbody>';
  tbl.innerHTML=hdr+body;
}}

/* ── 공통 라인 차트 옵션 ──
   Y축 min을 0보다 살짝 아래로 확장 → 0값 데이터 포인트가 X축 라벨과 겹치지 않음
   단, Y축 눈금 라벨에서는 음수 값을 표시하지 않음 (매출 데이터에 -가 자연스럽지 않음) */
function makeLineOpts(yLabel,yFmt,tooltipFmt){{
  return {{
    responsive:true,maintainAspectRatio:false,
    clip:false,
    layout:{{padding:{{top:22,bottom:8,left:4,right:4}}}},
    interaction:{{mode:'index',intersect:false}},
    plugins:{{
      legend:{{display:false}},
      tooltip:{{callbacks:{{label:ctx=>ctx.dataset.label+': '+(ctx.raw!==null?tooltipFmt(ctx.raw):'-')}}}}
    }},
    scales:{{
      x:{{ticks:{{font:{{size:11}}}},grid:{{display:false}}}},
      y:{{
        ticks:{{font:{{size:11}}, callback:v => v < 0 ? '' : yFmt(v)}},   // 음수 tick 라벨 숨김
        grid:{{display:false}},
        beginAtZero:true,
        afterDataLimits:(axis)=>{{
          const range = (axis.max - axis.min) || 1;
          axis.min = axis.min - range * 0.05;   // 0값 포인트가 X축과 겹치지 않도록 아래 5% 여유
        }}
      }}
    }}
  }};
}}

/* ── 범례 렌더 (선택 연도만큼 동적) ── */
function renderLegend(id,ys){{
  const el=document.getElementById(id); if(!el) return;
  el.innerHTML=ys.map(y=>`<span class="cl-item"><span class="cl-line" style="background:${{YEAR_COLORS[y]}}"></span>${{YEAR_LABELS[y]}}</span>`).join('');
}}

/* ── 차트 렌더 (선택 연도만큼 동적: 최대 2024/2025/2026/계획 4개 라인) ── */
function renderCharts(baseRows){{
  const ys=activeYears();
  renderLegend('legendRev',ys);
  renderLegend('legendQty',ys);
  renderLegend('legendCumRev',ys);
  renderLegend('legendCumQty',ys);

  // 앞뒤 빈 항목으로 1월/12월이 엣지에서 떨어지도록
  const labels=['','1월','2월','3월','4월','5월','6월','7월','8월','9월','10월','11월','12월',''];
  const tmpRev={{}}, tmpQty={{}}, hasMon={{}};
  ys.forEach(y=>{{ tmpRev[y]=new Array(12).fill(0); tmpQty[y]=new Array(12).fill(0); hasMon[y]=new Array(12).fill(false); }});
  baseRows.forEach(d=>{{
    const y=String(d.yr);
    if(!tmpRev[y]) return;
    const mi=d.month-1;
    tmpRev[y][mi]+=d.rev;
    tmpQty[y][mi]+=d.qty;
    if(d.rev||d.qty) hasMon[y][mi]=true;
  }});
  const revSeries={{}}, qtySeries={{}}, cumRevSeries={{}}, cumQtySeries={{}};
  ys.forEach(y=>{{
    const r=new Array(14).fill(null), q=new Array(14).fill(null);
    const cr=new Array(14).fill(null), cq=new Array(14).fill(null);
    // 데이터가 있는 마지막 월까지만 누적선 표시
    let lastMon=-1;
    for(let i=0;i<12;i++) if(hasMon[y][i]) lastMon=i;
    let cumR=0, cumQ=0;
    for(let i=0;i<12;i++){{
      r[i+1]=hasMon[y][i]?Math.round(tmpRev[y][i]/1e6):null;
      q[i+1]=hasMon[y][i]?Math.round(tmpQty[y][i]):null;
      cumR+=tmpRev[y][i]; cumQ+=tmpQty[y][i];
      if(i<=lastMon){{
        cr[i+1]=Math.round(cumR/1e6);
        cq[i+1]=Math.round(cumQ);
      }}
    }}
    revSeries[y]=r; qtySeries[y]=q; cumRevSeries[y]=cr; cumQtySeries[y]=cq;
  }});

  function mkDataset(y,data,thin){{
    return {{label:YEAR_LABELS[y],data,borderColor:YEAR_COLORS[y],
      backgroundColor:YEAR_COLORS[y]+'14',pointBackgroundColor:YEAR_COLORS[y],
      borderWidth:y==='2026'?2.5:2,pointRadius:thin?3:4,pointHoverRadius:thin?5:6,
      tension:0.3,spanGaps:false}};
  }}

  // 매출액 꺾은선 차트
  if(revChart) revChart.destroy();
  revChart=new Chart(document.getElementById('chartRev').getContext('2d'),{{
    type:'line',
    data:{{labels,datasets:ys.map(y=>mkDataset(y,revSeries[y],false))}},
    options:makeLineOpts('백만원',v=>v.toLocaleString(),v=>v.toLocaleString()+'백만원')
  }});

  // 누적 매출액 차트
  if(cumRevChart) cumRevChart.destroy();
  cumRevChart=new Chart(document.getElementById('chartCumRev').getContext('2d'),{{
    type:'line',
    data:{{labels,datasets:ys.map(y=>mkDataset(y,cumRevSeries[y],true))}},
    options:makeLineOpts('백만원',v=>v.toLocaleString(),v=>v.toLocaleString()+'백만원')
  }});

  // 누적 매출수량 차트
  if(cumQtyChart) cumQtyChart.destroy();
  cumQtyChart=new Chart(document.getElementById('chartCumQty').getContext('2d'),{{
    type:'line',
    data:{{labels,datasets:ys.map(y=>mkDataset(y,cumQtySeries[y],true))}},
    options:makeLineOpts('개',v=>v.toLocaleString(),v=>v.toLocaleString()+'개')
  }});

  // 수량 꺾은선 차트
  if(qtyChart) qtyChart.destroy();
  qtyChart=new Chart(document.getElementById('chartQty').getContext('2d'),{{
    type:'line',
    data:{{labels,datasets:ys.map(y=>mkDataset(y,qtySeries[y],true))}},
    options:makeLineOpts('개',v=>v.toLocaleString(),v=>v.toLocaleString()+'개')
  }});

  // ═══ Top5 채널 월별 매출 추이 ═══
  //   1) baseRows(월/분기 필터 적용 전)로 채널별 총 매출 계산 → 상위 5개 선정
  //   2) 상위 5개 채널의 월별 매출을 라인차트로 (선택 연도 중 최신 활성 연도 기준)
  const chTotals = new Map();
  baseRows.forEach(d => {{ if(d.channel) chTotals.set(d.channel, (chTotals.get(d.channel)||0) + d.rev); }});
  const top5 = [...chTotals.entries()].sort((a,b)=>b[1]-a[1]).slice(0,5).map(e => e[0]);
  const topYr = ys.includes('2026') ? '2026' : (ys.includes('2025') ? '2025' : ys[ys.length-1]);
  const chMonthly = new Map();  // ch → [12개월]
  top5.forEach(ch => chMonthly.set(ch, new Array(12).fill(0)));
  baseRows.forEach(d => {{
    if(String(d.yr) !== topYr) return;
    if(!chMonthly.has(d.channel)) return;
    chMonthly.get(d.channel)[d.month-1] += d.rev;
  }});
  const TOP5_COLORS = ['#1e3a8a','#dc2626','#059669','#d97706','#7c3aed'];
  const top5Datasets = top5.map((ch, i) => {{
    const monthly = chMonthly.get(ch);
    const data = new Array(14).fill(null);
    for(let m=0; m<12; m++) data[m+1] = monthly[m] ? Math.round(monthly[m]/1e6) : null;
    return {{
      label: ch,
      data: data,
      borderColor: TOP5_COLORS[i],
      backgroundColor: TOP5_COLORS[i]+'14',
      pointBackgroundColor: TOP5_COLORS[i],
      borderWidth: 2, pointRadius: 3, pointHoverRadius: 5,
      tension: 0.3, spanGaps: false
    }};
  }});
  // Top5 범례 렌더
  const legTop5 = document.getElementById('legendTop5');
  if(legTop5){{
    legTop5.innerHTML = top5.map((ch, i) =>
      `<span class="cl-item"><span class="cl-line" style="background:${{TOP5_COLORS[i]}}"></span>${{ch}}</span>`
    ).join('');
  }}
  if(top5Chart) top5Chart.destroy();
  top5Chart = new Chart(document.getElementById('chartTop5').getContext('2d'), {{
    type: 'line',
    data: {{ labels, datasets: top5Datasets }},
    options: makeLineOpts('백만원', v=>v.toLocaleString(), v=>v.toLocaleString()+'백만원')
  }});
  // 차트 제목에 기준 연도 표시 (씨엠에스랩 로고 색상 유지)
  const top5Title = document.querySelector('#chartTop5Card .chart-card-title');
  if(top5Title) top5Title.textContent = `Top5 채널 월별 매출 추이 (${{topYr}}년 · 백만원)`;
  // Top5 상세 데이터 표 갱신
  renderTop5Table(top5, chMonthly);

  renderChartTables(baseRows);
}}

/* ── 표 렌더 공용 헬퍼 ──
   컬럼 폭 % 배분:
     · 라벨 열: 30% (계층 깊은 이름 넉넉하게)
     · 연도 그룹(25/26) + 증감 그룹 → 각 그룹이 동일한 전체 폭
     · 각 그룹 내부: 수량 40 / 매출액 30 / GP율 30 (모든 그룹에서 동일 비율)
   → 결과적으로 25년/26년/증감의 수량 열들끼리, 매출액 열들끼리, GP율 열들끼리 모두 폭이 같음
*/
function hasDiffCol(yrs){{ return yrs.indexOf('2025') >= 0 && yrs.indexOf('2026') >= 0; }}

function makeColgroup(yrs){{
  const hasDiff = hasDiffCol(yrs);
  const fixPct  = 34;   // 라벨 30 → 34% (계층 이름 더 넉넉하게, 25/26/증감 폭 축소)
  const groupCount = yrs.length + (hasDiff ? 1 : 0);
  const perGroup = (100 - fixPct) / groupCount;
  // 모든 그룹(연도·증감)에서 동일 비율: 수량 40% / 매출액 30% / GP율 30%
  const qCol = (perGroup * 40/100).toFixed(3);
  const rCol = (perGroup * 30/100).toFixed(3);
  const gCol = (perGroup * 30/100).toFixed(3);
  let cols = '<col style="width:'+fixPct+'%">';
  yrs.forEach(y => {{
    cols += '<col style="width:'+qCol+'%">';
    cols += '<col style="width:'+rCol+'%">';
    cols += '<col style="width:'+gCol+'%">';
  }});
  if(hasDiff){{
    cols += '<col style="width:'+qCol+'%">';
    cols += '<col style="width:'+rCol+'%">';
    cols += '<col style="width:'+gCol+'%">';
  }}
  return '<colgroup>' + cols + '</colgroup>';
}}

/* 증감 값 포맷 & 색상 클래스. 양수는 '+' 부호, 음수는 회계 관례대로 괄호로 표기 */
function fmtQtyDiff(diff){{
  const v = Math.round(diff);
  if(v === 0) return '-';
  return v > 0 ? '+' + v.toLocaleString() : '(' + Math.abs(v).toLocaleString() + ')';
}}
function fmtRevDiff(diff){{
  const v = Math.round(diff/1e6);
  if(v === 0) return '-';
  return v > 0 ? '+' + v.toLocaleString() : '(' + Math.abs(v).toLocaleString() + ')';
}}
function fmtGpDiff(t25, t26){{
  if(!t25.rev || !t26.rev) return '-';
  const gp25 = (t25.rev-t25.cost)/t25.rev*100;
  const gp26 = (t26.rev-t26.cost)/t26.rev*100;
  const d = gp26 - gp25;
  if(Math.abs(d) < 0.05) return '-';
  const abs = Math.abs(d).toFixed(1);
  return d > 0 ? '+' + abs + '%p' : '(' + abs + '%p)';
}}
function diffCls(diff){{
  if(!diff || Math.abs(diff) < 0.5) return 'diff-neu';
  return diff > 0 ? 'diff-pos' : 'diff-neg';
}}
function gpDiffCls(t25, t26){{
  if(!t25.rev || !t26.rev) return 'diff-neu';
  const d = (t26.rev-t26.cost)/t26.rev*100 - (t25.rev-t25.cost)/t25.rev*100;
  if(Math.abs(d) < 0.05) return 'diff-neu';
  return d > 0 ? 'diff-pos' : 'diff-neg';
}}

/* 증감 셀 3개 반환 (수량/매출액/GP%p). yrs에 25/26 둘 다 없으면 빈 문자열 */
function makeDiffCells(yrs, t){{
  if(!hasDiffCol(yrs)) return '';
  const t25 = t['2025'] || {{qty:0,rev:0,cost:0}};
  const t26 = t['2026'] || {{qty:0,rev:0,cost:0}};
  const qD = t26.qty - t25.qty;
  const rD = t26.rev - t25.rev;
  return `<td class="yc-diff y-sep-l ${{diffCls(qD)}}">${{fmtQtyDiff(qD)}}</td>`
       + `<td class="yc-diff ${{diffCls(rD)}}">${{fmtRevDiff(rD)}}</td>`
       + `<td class="yc-diff ${{gpDiffCls(t25,t26)}}">${{fmtGpDiff(t25,t26)}}</td>`;
}}

/* ── 표 렌더 ── */
function renderTable(rows){{
  lastRows=rows;
  const q=document.getElementById('tblSearch').value.trim().toLowerCase();

  // yr 목록 정렬
  const yrs=activeYears();

  // 품목군→품목분류→SKU 집계 (SKU엔 sale_type도 유지)
  const grpMap=new Map();
  rows.forEach(d=>{{
    const g=d.item_group||'(미분류)', c=d.item_cat||'(미분류)', s=d.sku, sn=d.sku_name||d.sku;
    if(!grpMap.has(g)) grpMap.set(g,new Map());
    const catMap=grpMap.get(g);
    if(!catMap.has(c)) catMap.set(c,new Map());
    const skuMap=catMap.get(c);
    const key=s+'|'+sn;
    if(!skuMap.has(key)) skuMap.set(key,{{sku:s,sku_name:sn,sale_type:d.sale_type,data:{{}},channels:new Map()}});
    const entry=skuMap.get(key);
    const yr=String(d.yr);
    if(!entry.data[yr]) entry.data[yr]={{qty:0,rev:0,cost:0}};
    entry.data[yr].qty+=d.qty; entry.data[yr].rev+=d.rev; entry.data[yr].cost+=d.cost;
    // 채널(영업그룹)별 세분화 — SKU 하위 레벨용
    const ch=d.channel||'(미분류)';
    if(!entry.channels.has(ch)) entry.channels.set(ch,{{}});
    const cd=entry.channels.get(ch);
    if(!cd[yr]) cd[yr]={{qty:0,rev:0,cost:0}};
    cd[yr].qty+=d.qty; cd[yr].rev+=d.rev; cd[yr].cost+=d.cost;
  }});

  // 헤더 구성 — 각 연도 열에 yh-YYYY, 우측에 yh-diff(증감) 배치
  const hasDiff = hasDiffCol(yrs);
  let yrCols='';
  yrs.forEach(y=>{{ yrCols+=`<th class="yh-${{y}}" colspan="3">${{YEAR_LABELS[y]}}</th>`; }});
  if(hasDiff) yrCols += `<th class="yh-diff" colspan="3">증감</th>`;
  let subCols='';
  yrs.forEach(y=>{{
    subCols+=`<th class="ys-${{y}} y-sep-l" title="${{YEAR_LABELS[y]}} 수량">수량<span class="unit">(개)</span></th>`
          + `<th class="ys-${{y}}" title="${{YEAR_LABELS[y]}} 매출액">매출액<span class="unit">(백만원)</span></th>`
          + `<th class="ys-${{y}}" title="${{YEAR_LABELS[y]}} GP율">매출총이익률<span class="unit">(%)</span></th>`;
  }});
  if(hasDiff){{
    subCols += `<th class="ys-diff y-sep-l" title="증감 수량">수량<span class="unit">(개)</span></th>`
             + `<th class="ys-diff" title="증감 매출액">매출액<span class="unit">(백만원)</span></th>`
             + `<th class="ys-diff" title="매출총이익률 증감">매출총이익률<span class="unit">(%p)</span></th>`;
  }}
  const colgroup = makeColgroup(yrs);

  let html=`<table>${{colgroup}}<thead>
  <tr><th class="fix-col" rowspan="2">품목군 / 품목분류 / SKU / 채널</th>${{yrCols}}</tr>
  <tr>${{subCols}}</tr>
  </thead><tbody>`;

  let skuCount=0;
  const grpSub=[];   // gi → 하위(품목분류/SKU/채널) HTML : 펼칠 때 지연 삽입
  // GRP_ORDER(25년 매출액 내림차순)에 따라 정렬, 목록에 없는 그룹은 뒤에 추가
  const orderedGrps=GRP_ORDER.filter(g=>grpMap.has(g));
  const extraGrps=[...grpMap.keys()].filter(g=>!GRP_ORDER.includes(g)).sort();
  const grpKeys=[...orderedGrps,...extraGrps];
  grpKeys.forEach((g,gi)=>{{
    if(q && !g.toLowerCase().includes(q)){{
      // 하위에 매치가 없으면 스킵 (아래에서 체크)
    }}
    const catMap=grpMap.get(g);
    // 품목군 합계
    const grpTot={{}};
    yrs.forEach(y=>grpTot[y]={{qty:0,rev:0,cost:0}});
    catMap.forEach((skuMap,c)=>skuMap.forEach(e=>yrs.forEach(y=>{{
      if(e.data[y]){{ grpTot[y].qty+=e.data[y].qty; grpTot[y].rev+=e.data[y].rev; grpTot[y].cost+=e.data[y].cost; }}
    }})));
    let grpCells=yrs.map(y=>{{
      const t=grpTot[y];
      const cr=fmtCR(t.rev,t.cost);
      return `<td class="yc-${{y}} y-sep-l">${{fmtQ(t.qty)}}</td><td class="yc-${{y}}">${{fmtRev(t.rev)}}</td><td class="yc-${{y}}">${{cr}}</td>`;
    }}).join('') + makeDiffCells(yrs, grpTot);

    let hasMatch=false;
    const catKeys=[...catMap.keys()].sort((a,b)=>{{
      if(a==='(미분류)') return 1;
      if(b==='(미분류)') return -1;
      return a.localeCompare(b,'ko');
    }});
    let catHtml='';
    catKeys.forEach((c,ci)=>{{
      const skuMap=catMap.get(c);
      const catTot={{}};
      yrs.forEach(y=>catTot[y]={{qty:0,rev:0,cost:0}});
      skuMap.forEach(e=>yrs.forEach(y=>{{
        if(e.data[y]){{ catTot[y].qty+=e.data[y].qty; catTot[y].rev+=e.data[y].rev; catTot[y].cost+=e.data[y].cost; }}
      }}));
      let catCells=yrs.map(y=>{{
        const t=catTot[y];
        const cr=fmtCR(t.rev,t.cost);
        return `<td class="yc-${{y}} y-sep-l">${{fmtQ(t.qty)}}</td><td class="yc-${{y}}">${{fmtRev(t.rev)}}</td><td class="yc-${{y}}">${{cr}}</td>`;
      }}).join('') + makeDiffCells(yrs, catTot);

      let skuHtml='', skuMatch=false;
      let skuIdx=0;   // 카테고리 내 SKU 인덱스 (팀 서브행 그룹핑 키)
      skuMap.forEach(e=>{{
        const lbl=(e.sku+' '+e.sku_name).toLowerCase();
        if(q && !lbl.includes(q) && !c.toLowerCase().includes(q) && !g.toLowerCase().includes(q)) return;
        skuMatch=true; skuCount++;
        const si=skuIdx++;
        let skuCells=yrs.map(y=>{{
          const t=e.data[y]||{{qty:0,rev:0,cost:0}};
          const cr=fmtCR(t.rev,t.cost);
          return `<td class="yc-${{y}} y-sep-l">${{fmtQ(t.qty)}}</td><td class="yc-${{y}}">${{fmtRev(t.rev)}}</td><td class="yc-${{y}}">${{cr}}</td>`;
        }}).join('') + makeDiffCells(yrs, e.data);
        // 판매품/증정품 배지 (SKU# 앞에 위치)
        const stCls = e.sale_type==='판매품'?'sb-sale':e.sale_type==='증정품'?'sb-gift':'sb-none';
        const stTxt = e.sale_type||'-';
        // 채널 존재하면 SKU 행에 토글 화살표 부여
        const hasChans = e.channels && e.channels.size>0;
        const arrow = hasChans ? `<span class="toggle-arrow" onclick="toggleSku(${{gi}},${{ci}},${{si}},this)" title="채널 펼치기/접기">▸</span>` : '';
        skuHtml+=`<tr class="sku-row" data-g="${{gi}}" data-c="${{ci}}" data-si="${{si}}"><td class="fix-col" title="${{e.sku_name}}">${{arrow}}<span class="sale-badge ${{stCls}}">${{stTxt}}</span>${{e.sku}} ${{e.sku_name}}</td>${{skuCells}}</tr>`;
        // 채널별 서브 행 (기본 hidden) — SKU 하위 레벨
        if(hasChans){{
          const chanNames=[...e.channels.keys()].sort((a,b)=>a.localeCompare(b,'ko'));
          chanNames.forEach(ch=>{{
            const chanData=e.channels.get(ch);
            let chanCells=yrs.map(y=>{{
              const t=chanData[y]||{{qty:0,rev:0,cost:0}};
              const cr=fmtCR(t.rev,t.cost);
              return `<td class="yc-${{y}} y-sep-l">${{fmtQ(t.qty)}}</td><td class="yc-${{y}}">${{fmtRev(t.rev)}}</td><td class="yc-${{y}}">${{cr}}</td>`;
            }}).join('') + makeDiffCells(yrs, chanData);
            skuHtml+=`<tr class="chan-row hidden" data-g="${{gi}}" data-c="${{ci}}" data-si="${{si}}"><td class="fix-col">└ ${{ch}}</td>${{chanCells}}</tr>`;
          }});
        }}
      }});

      if(q && !skuMatch && !c.toLowerCase().includes(q) && !g.toLowerCase().includes(q)) return;
      hasMatch=true;
      // cat-row: 기본 hidden, 화살표 span만 onclick
      catHtml+=`<tr class="cat-row hidden" data-g="${{gi}}" data-c="${{ci}}"><td class="fix-col"><span class="toggle-arrow" onclick="toggleCat(${{gi}},${{ci}},this)" title="SKU 펼치기/접기">▸</span>${{c}}</td>${{catCells}}</tr>`;
      // sku-row: 기본 hidden
      catHtml+=skuHtml.replace(/class="sku-row"/g,'class="sku-row hidden"');
    }});

    if(!hasMatch && q) return;
    // grp-row: 화살표 span만 onclick
    html+=`<tr class="grp-row" data-g="${{gi}}"><td class="fix-col"><span class="toggle-arrow" onclick="toggleGrpRows(${{gi}},this)" title="품목분류 펼치기/접기">▸</span> ${{g}}</td>${{grpCells}}</tr>`;
    grpSub[gi]=catHtml;        // 하위는 펼칠 때 삽입(지연 렌더)
    if(q) html+=catHtml;       // 검색 중엔 즉시 렌더(기존 동작 유지)
  }});
  window.__t1Sub=grpSub;

  // ═══ 합계 행 (필터/검색이 반영된 rows 전체 총합) ═══
  const grandTot={{}};
  yrs.forEach(y=>grandTot[y]={{qty:0,rev:0,cost:0}});
  rows.forEach(d=>{{
    const y=String(d.yr);
    if(grandTot[y]){{ grandTot[y].qty+=d.qty; grandTot[y].rev+=d.rev; grandTot[y].cost+=d.cost; }}
  }});
  const totalCells = yrs.map(y=>{{
    const t=grandTot[y]; const cr=fmtCR(t.rev,t.cost);
    return `<td class="yc-${{y}} y-sep-l">${{fmtQ(t.qty)}}</td><td class="yc-${{y}}">${{fmtRev(t.rev)}}</td><td class="yc-${{y}}">${{cr}}</td>`;
  }}).join('') + makeDiffCells(yrs, grandTot);
  html += `<tr class="total-row"><td class="fix-col">합계</td>${{totalCells}}</tr>`;

  html+='</tbody></table>';
  document.getElementById('tblOuter').innerHTML=html;
  document.getElementById('tblCount').textContent=`SKU ${{skuCount.toLocaleString()}}건`;
}}

/* ── 표 토글 ── */
function toggleGrpRows(gi,arrowEl){{
  let catRows=[...document.querySelectorAll('.cat-row[data-g="'+gi+'"]')];
  // 지연 생성: 하위 행이 아직 없으면 저장해둔 HTML을 품목군 행 뒤에 삽입
  if(catRows.length===0 && window.__t1Sub && window.__t1Sub[gi]){{
    const gRow=document.querySelector('#tblOuter tr.grp-row[data-g="'+gi+'"]');
    if(gRow){{ gRow.insertAdjacentHTML('afterend', window.__t1Sub[gi]); catRows=[...document.querySelectorAll('.cat-row[data-g="'+gi+'"]')]; }}
  }}
  const hiding=catRows.some(r=>!r.classList.contains('hidden'));
  catRows.forEach(r=>r.classList.toggle('hidden',hiding));
  // 접을 때 하위 SKU/팀 행도 함께 접기 + 화살표 리셋
  if(hiding){{
    document.querySelectorAll('.sku-row[data-g="'+gi+'"]').forEach(r=>r.classList.add('hidden'));
    document.querySelectorAll('.chan-row[data-g="'+gi+'"]').forEach(r=>r.classList.add('hidden'));
    document.querySelectorAll('.cat-row[data-g="'+gi+'"] .toggle-arrow').forEach(a=>a.textContent='▸');
    document.querySelectorAll('.sku-row[data-g="'+gi+'"] .toggle-arrow').forEach(a=>a.textContent='▸');
  }}
  arrowEl.textContent=hiding?'▸':'▾';
}}
function toggleCat(gi,ci,arrowEl){{
  const skuRows=document.querySelectorAll('.sku-row[data-g="'+gi+'"][data-c="'+ci+'"]');
  const hiding=skuRows.length&&!skuRows[0].classList.contains('hidden');
  skuRows.forEach(r=>r.classList.toggle('hidden',hiding));
  // 접을 때 하위 팀 행도 함께 접기 + SKU 화살표 리셋
  if(hiding){{
    document.querySelectorAll('.chan-row[data-g="'+gi+'"][data-c="'+ci+'"]').forEach(r=>r.classList.add('hidden'));
    document.querySelectorAll('.sku-row[data-g="'+gi+'"][data-c="'+ci+'"] .toggle-arrow').forEach(a=>a.textContent='▸');
  }}
  arrowEl.textContent=hiding?'▸':'▾';
}}
function toggleSku(gi,ci,si,arrowEl){{
  const chanRows=document.querySelectorAll('.chan-row[data-g="'+gi+'"][data-c="'+ci+'"][data-si="'+si+'"]');
  const hiding=chanRows.length && !chanRows[0].classList.contains('hidden');
  chanRows.forEach(r=>r.classList.toggle('hidden',hiding));
  arrowEl.textContent=hiding?'▸':'▾';
}}

/* 품목별 매출 현황 표 전체 펼치기 (품목분류까지 · SKU/채널은 접힘) */
function expandAllTbl1(){{
  // 지연 생성분 먼저 모두 삽입
  if(window.__t1Sub){{
    document.querySelectorAll('#tblOuter tr.grp-row').forEach(gRow => {{
      const gi=gRow.getAttribute('data-g');
      if(document.querySelectorAll('.cat-row[data-g="'+gi+'"]').length===0 && window.__t1Sub[gi]){{
        gRow.insertAdjacentHTML('afterend', window.__t1Sub[gi]);
      }}
    }});
  }}
  document.querySelectorAll('#tblOuter .cat-row').forEach(r => r.classList.remove('hidden'));
  document.querySelectorAll('#tblOuter .grp-row .toggle-arrow').forEach(a => a.textContent = '▾');
  // SKU, 채널 행은 hidden 유지 + 화살표 리셋
  document.querySelectorAll('#tblOuter .sku-row').forEach(r => r.classList.add('hidden'));
  document.querySelectorAll('#tblOuter .chan-row').forEach(r => r.classList.add('hidden'));
  document.querySelectorAll('#tblOuter .cat-row .toggle-arrow').forEach(a => a.textContent = '▸');
  document.querySelectorAll('#tblOuter .sku-row .toggle-arrow').forEach(a => a.textContent = '▸');
}}
function collapseAllTbl1(){{
  document.querySelectorAll('#tblOuter .cat-row').forEach(r => r.classList.add('hidden'));
  document.querySelectorAll('#tblOuter .sku-row').forEach(r => r.classList.add('hidden'));
  document.querySelectorAll('#tblOuter .chan-row').forEach(r => r.classList.add('hidden'));
  document.querySelectorAll('#tblOuter .toggle-arrow').forEach(a => a.textContent = '▸');
}}

/* 채널별 매출 현황 표 전체 펼치기 (채널 → 거래처 → 품목군 → 품목분류까지) */
function expandAllTbl2(){{
  // 지연 생성분 먼저 모두 생성
  if(window.__t2Built && window.__t2BuildSub){{
    document.querySelectorAll('#tblOuter2 tr.n2-ch').forEach(chRow => {{
      const chi = chRow.getAttribute('data-ch');
      if(!window.__t2Built[chi]){{ chRow.insertAdjacentHTML('afterend', window.__t2BuildSub(chi)); window.__t2Built[chi] = true; }}
    }});
  }}
  document.querySelectorAll('#tblOuter2 .n2-cs').forEach(r => r.classList.remove('hidden'));
  document.querySelectorAll('#tblOuter2 .n2-gp').forEach(r => r.classList.remove('hidden'));
  document.querySelectorAll('#tblOuter2 .n2-ct').forEach(r => r.classList.remove('hidden'));
  document.querySelectorAll('#tblOuter2 .n2-sk').forEach(r => r.classList.add('hidden'));
  // 화살표 상태: 채널/거래처/품목군 → ▾, 품목분류 → ▸ (하위 SKU 접힘)
  document.querySelectorAll('#tblOuter2 .n2-ch .toggle-arrow').forEach(a => a.textContent = '▾');
  document.querySelectorAll('#tblOuter2 .n2-cs .toggle-arrow').forEach(a => a.textContent = '▾');
  document.querySelectorAll('#tblOuter2 .n2-gp .toggle-arrow').forEach(a => a.textContent = '▾');
  document.querySelectorAll('#tblOuter2 .n2-ct .toggle-arrow').forEach(a => a.textContent = '▸');
}}
function collapseAllTbl2(){{
  document.querySelectorAll('#tblOuter2 .n2-cs').forEach(r => r.classList.add('hidden'));
  document.querySelectorAll('#tblOuter2 .n2-gp').forEach(r => r.classList.add('hidden'));
  document.querySelectorAll('#tblOuter2 .n2-ct').forEach(r => r.classList.add('hidden'));
  document.querySelectorAll('#tblOuter2 .n2-sk').forEach(r => r.classList.add('hidden'));
  document.querySelectorAll('#tblOuter2 .toggle-arrow').forEach(a => a.textContent = '▸');
}}

/* ══════════════════ 두 번째 표 (채널>거래처>품목군>품목분류>SKU) 렌더 ══════════════════ */
function renderTable2(rows){{
  const q = (document.getElementById('tblSearch2')?.value||'').trim().toLowerCase();
  const yrs = activeYears();

  // 5단계 중첩 Map
  const chMap = new Map();
  rows.forEach(d => {{
    const ch = d.channel  || '(미분류)';
    const cs = d.customer || '(미분류)';
    const gp = d.item_group || '(미분류)';
    const ct = d.item_cat   || '(미분류)';
    const sk = d.sku, sn = d.sku_name || d.sku;
    if(!chMap.has(ch)) chMap.set(ch, new Map());
    const csMap = chMap.get(ch);
    if(!csMap.has(cs)) csMap.set(cs, new Map());
    const gpMap = csMap.get(cs);
    if(!gpMap.has(gp)) gpMap.set(gp, new Map());
    const ctMap = gpMap.get(gp);
    if(!ctMap.has(ct)) ctMap.set(ct, new Map());
    const skMap = ctMap.get(ct);
    const key = sk + '|' + sn;
    if(!skMap.has(key)) skMap.set(key, {{sku:sk, sku_name:sn, sale_type:d.sale_type, data:{{}}}});
    const entry = skMap.get(key);
    const yr = String(d.yr);
    if(!entry.data[yr]) entry.data[yr] = {{qty:0,rev:0,cost:0}};
    entry.data[yr].qty+=d.qty; entry.data[yr].rev+=d.rev; entry.data[yr].cost+=d.cost;
  }});

  // 헤더 — 상단 표와 동일 구조 (25/26 활성이면 증감 컬럼 추가)
  const hasDiff = hasDiffCol(yrs);
  let yrCols='';
  yrs.forEach(y => {{ yrCols += `<th class="yh-${{y}}" colspan="3">${{YEAR_LABELS[y]}}</th>`; }});
  if(hasDiff) yrCols += `<th class="yh-diff" colspan="3">증감</th>`;
  let subCols='';
  yrs.forEach(y => {{
    subCols += `<th class="ys-${{y}} y-sep-l">수량<span class="unit">(개)</span></th>`
            +  `<th class="ys-${{y}}">매출액<span class="unit">(백만원)</span></th>`
            +  `<th class="ys-${{y}}">매출총이익률<span class="unit">(%)</span></th>`;
  }});
  if(hasDiff){{
    subCols += `<th class="ys-diff y-sep-l">수량<span class="unit">(개)</span></th>`
            +  `<th class="ys-diff">매출액<span class="unit">(백만원)</span></th>`
            +  `<th class="ys-diff">매출총이익률<span class="unit">(%p)</span></th>`;
  }}
  const colgroup = makeColgroup(yrs);

  let html = `<table>${{colgroup}}<thead>
  <tr><th class="fix-col" rowspan="2">채널 / 거래처 / 품목군 / 품목분류 / SKU</th>${{yrCols}}</tr>
  <tr>${{subCols}}</tr>
  </thead><tbody>`;

  // 셀 렌더 헬퍼 — 연도별 3셀 + 증감 3셀(활성 시)
  function makeCells(t){{
    const yrCells = yrs.map(y => {{
      const b = t[y] || {{qty:0,rev:0,cost:0}};
      const cr = fmtCR(b.rev, b.cost);
      return `<td class="yc-${{y}} y-sep-l">${{fmtQ(b.qty)}}</td><td class="yc-${{y}}">${{fmtRev(b.rev)}}</td><td class="yc-${{y}}">${{cr}}</td>`;
    }}).join('');
    return yrCells + makeDiffCells(yrs, t);
  }}
  function newTot(){{ const t={{}}; yrs.forEach(y=>t[y]={{qty:0,rev:0,cost:0}}); return t; }}
  function addTot(dst, src){{
    yrs.forEach(y=>{{ if(src[y]){{ dst[y].qty+=src[y].qty; dst[y].rev+=src[y].rev; dst[y].cost+=src[y].cost; }} }});
  }}

  // 채널명 정렬: 팀참고 파일 순서 우선, 나머지는 한글순
  const chNames = sortByChannelOrder([...chMap.keys()]);
  let skuCount = 0;
  const chStore = [];   // chi → {{csAggs}} : 하위행 지연(lazy) 생성용 데이터 보관

  // 채널 하위(거래처>품목군>품목분류>SKU) 행 HTML 생성. makeCells 재사용 → 값 100% 동일.
  // 초기엔 채널행만 그리고, 펼칠 때 이 함수로 하위 행을 생성한다(클릭 렉 방지).
  function buildSub(chi){{
    const st = chStore[chi];
    if(!st) return '';
    const csAggs = st.csAggs;
    const csKeys = [...csAggs.keys()];
    let h = '';
    csAggs.forEach((csD, cs) => {{
      const csi = csKeys.indexOf(cs);
      h += `<tr class="n2-cs hidden" data-ch="${{chi}}" data-cs="${{csi}}"><td class="fix-col"><span class="toggle-arrow" onclick="tog2('cs',${{chi}},${{csi}},0,0,this)" title="펼치기/접기">▸</span> ${{cs}}</td>${{makeCells(csD.csTot)}}</tr>`;
      const gpKeys = [...csD.gpAggs.keys()];
      csD.gpAggs.forEach((gpD, gp) => {{
        const gpi = gpKeys.indexOf(gp);
        h += `<tr class="n2-gp hidden" data-ch="${{chi}}" data-cs="${{csi}}" data-gp="${{gpi}}"><td class="fix-col"><span class="toggle-arrow" onclick="tog2('gp',${{chi}},${{csi}},${{gpi}},0,this)" title="펼치기/접기">▸</span> ${{gp}}</td>${{makeCells(gpD.gpTot)}}</tr>`;
        const ctKeys = [...gpD.ctAggs.keys()];
        gpD.ctAggs.forEach((ctD, ct) => {{
          const cti = ctKeys.indexOf(ct);
          h += `<tr class="n2-ct hidden" data-ch="${{chi}}" data-cs="${{csi}}" data-gp="${{gpi}}" data-ct="${{cti}}"><td class="fix-col"><span class="toggle-arrow" onclick="tog2('ct',${{chi}},${{csi}},${{gpi}},${{cti}},this)" title="펼치기/접기">▸</span> ${{ct}}</td>${{makeCells(ctD.ctTot)}}</tr>`;
          ctD.skus.forEach(sd => {{
            const stCls = sd.e.sale_type==='판매품'?'sb-sale':sd.e.sale_type==='증정품'?'sb-gift':'sb-none';
            const stTxt = sd.e.sale_type||'-';
            h += `<tr class="n2-sk hidden" data-ch="${{chi}}" data-cs="${{csi}}" data-gp="${{gpi}}" data-ct="${{cti}}"><td class="fix-col" title="${{sd.e.sku_name}}"><span class="sale-badge ${{stCls}}">${{stTxt}}</span>${{sd.e.sku}} ${{sd.e.sku_name}}</td>${{makeCells(sd.totForSku)}}</tr>`;
          }});
        }});
      }});
    }});
    return h;
  }}
  window.__t2BuildSub = buildSub;
  window.__t2Built = {{}};

  chNames.forEach((ch, chi) => {{
    const csMap = chMap.get(ch);
    // 채널 합계 = 하위 전체
    const chTot = newTot();
    let chSku = 0;
    // 거래처 별로 진행, 하위 합계 계산
    const csNames = [...csMap.keys()].sort((a,b)=>a.localeCompare(b,'ko'));

    // 우선 채널/거래처/품목군/분류/SKU 별 raw 집계 후 채널 총합 계산
    // (실시간 계산이지만 O(n)이라 부담 없음)
    const csAggs = new Map();  // cs → {{tot, gpAggs}}
    csNames.forEach((cs, csi) => {{
      const gpMap = csMap.get(cs);
      const csTot = newTot();
      const gpNames = [...gpMap.keys()].sort((a,b)=>a.localeCompare(b,'ko'));
      const gpAggs = new Map();
      gpNames.forEach((gp, gpi) => {{
        const ctMap = gpMap.get(gp);
        const gpTot = newTot();
        const ctNames = [...ctMap.keys()].sort((a,b)=>{{
          if(a==='(미분류)') return 1; if(b==='(미분류)') return -1;
          return a.localeCompare(b,'ko');
        }});
        const ctAggs = new Map();
        ctNames.forEach((ct, cti) => {{
          const skMap = ctMap.get(ct);
          const ctTot = newTot();
          const skus = [];
          skMap.forEach(e => {{
            const totForSku = newTot();
            addTot(totForSku, e.data);
            addTot(ctTot, e.data);
            skus.push({{e, totForSku}});
          }});
          chSku += skus.length;
          addTot(gpTot, ctTot);
          ctAggs.set(ct, {{ctTot, skus}});
        }});
        addTot(csTot, gpTot);
        gpAggs.set(gp, {{gpTot, ctAggs, ctNames}});
      }});
      addTot(chTot, csTot);
      csAggs.set(cs, {{csTot, gpAggs, gpNames}});
    }});
    chStore[chi] = {{csAggs}};

    // 검색어 필터: 매치 없으면 채널 전체 스킵
    let chHasMatch = false;
    if(q){{
      if(ch.toLowerCase().includes(q)) chHasMatch = true;
      csAggs.forEach((csD, cs)=>{{
        if(cs.toLowerCase().includes(q)) chHasMatch = true;
        csD.gpAggs.forEach((gpD, gp)=>{{
          if(gp.toLowerCase().includes(q)) chHasMatch = true;
          gpD.ctAggs.forEach((ctD, ct)=>{{
            if(ct.toLowerCase().includes(q)) chHasMatch = true;
            ctD.skus.forEach(sd=>{{
              if((sd.e.sku+' '+sd.e.sku_name).toLowerCase().includes(q)) chHasMatch = true;
            }});
          }});
        }});
      }});
      if(!chHasMatch) return;
    }}

    skuCount += chSku;   // 표시되는 채널의 SKU만 카운트(검색 스킵 채널 제외 — 기존 동작 동일)
    // 채널 행 (하위 거래처/품목군/품목분류/SKU는 펼칠 때 buildSub로 지연 생성)
    html += `<tr class="n2-ch" data-ch="${{chi}}"><td class="fix-col"><span class="toggle-arrow" onclick="tog2('ch',${{chi}},0,0,0,this)" title="펼치기/접기">▸</span> ${{ch}}</td>${{makeCells(chTot)}}</tr>`;
    // 검색 중이면 하위를 즉시 생성(기존 검색 동작 유지)
    if(q){{ html += buildSub(chi); window.__t2Built[chi] = true; }}
  }});

  // ═══ 합계 행 (필터/검색이 반영된 rows 전체 총합) ═══
  const grandTot2={{}};
  yrs.forEach(y=>grandTot2[y]={{qty:0,rev:0,cost:0}});
  rows.forEach(d=>{{
    const y=String(d.yr);
    if(grandTot2[y]){{ grandTot2[y].qty+=d.qty; grandTot2[y].rev+=d.rev; grandTot2[y].cost+=d.cost; }}
  }});
  html += `<tr class="total-row"><td class="fix-col">합계</td>${{makeCells(grandTot2)}}</tr>`;

  html += '</tbody></table>';
  document.getElementById('tblOuter2').innerHTML = html;
  document.getElementById('tblCount2').textContent = `SKU ${{skuCount.toLocaleString()}}건`;
}}

/* 두 번째 표의 5단계 토글 (level: ch/cs/gp/ct) */
function tog2(level, chi, csi, gpi, cti, arrowEl){{
  // 지연 생성: 채널을 처음 펼칠 때 하위 행을 생성해 채널행 뒤에 삽입
  if(level === 'ch' && window.__t2Built && !window.__t2Built[chi] && window.__t2BuildSub){{
    const chRow = document.querySelector('#tblOuter2 tr.n2-ch[data-ch="'+chi+'"]');
    if(chRow){{ chRow.insertAdjacentHTML('afterend', window.__t2BuildSub(chi)); window.__t2Built[chi] = true; }}
  }}
  let sel = '';
  if(level === 'ch') sel = 'tr.n2-cs[data-ch="'+chi+'"]';
  else if(level === 'cs') sel = 'tr.n2-gp[data-ch="'+chi+'"][data-cs="'+csi+'"]';
  else if(level === 'gp') sel = 'tr.n2-ct[data-ch="'+chi+'"][data-cs="'+csi+'"][data-gp="'+gpi+'"]';
  else if(level === 'ct') sel = 'tr.n2-sk[data-ch="'+chi+'"][data-cs="'+csi+'"][data-gp="'+gpi+'"][data-ct="'+cti+'"]';
  const rows = document.querySelectorAll(sel);
  const hiding = rows.length && !rows[0].classList.contains('hidden');
  rows.forEach(r => r.classList.toggle('hidden', hiding));
  // 접을 때 하위 레벨도 모두 접기
  if(hiding){{
    let deeperSels = [];
    if(level === 'ch') deeperSels = [
      'tr.n2-gp[data-ch="'+chi+'"]',
      'tr.n2-ct[data-ch="'+chi+'"]',
      'tr.n2-sk[data-ch="'+chi+'"]',
    ];
    else if(level === 'cs') deeperSels = [
      'tr.n2-ct[data-ch="'+chi+'"][data-cs="'+csi+'"]',
      'tr.n2-sk[data-ch="'+chi+'"][data-cs="'+csi+'"]',
    ];
    else if(level === 'gp') deeperSels = [
      'tr.n2-sk[data-ch="'+chi+'"][data-cs="'+csi+'"][data-gp="'+gpi+'"]',
    ];
    deeperSels.forEach(s => document.querySelectorAll(s).forEach(r => r.classList.add('hidden')));
    // 하위 화살표 리셋
    let arrowSels = [];
    if(level === 'ch') arrowSels = ['tr.n2-cs[data-ch="'+chi+'"] .toggle-arrow','tr.n2-gp[data-ch="'+chi+'"] .toggle-arrow','tr.n2-ct[data-ch="'+chi+'"] .toggle-arrow'];
    else if(level === 'cs') arrowSels = ['tr.n2-gp[data-ch="'+chi+'"][data-cs="'+csi+'"] .toggle-arrow','tr.n2-ct[data-ch="'+chi+'"][data-cs="'+csi+'"] .toggle-arrow'];
    else if(level === 'gp') arrowSels = ['tr.n2-ct[data-ch="'+chi+'"][data-cs="'+csi+'"][data-gp="'+gpi+'"] .toggle-arrow'];
    arrowSels.forEach(s => document.querySelectorAll(s).forEach(a => a.textContent='▸'));
  }}
  arrowEl.textContent = hiding ? '▸' : '▾';
}}

/* ── 메인 렌더 ── */
/* 사이드바 visibility DOM 업데이트 (activeGrps/Cats는 render()에서 함께 계산해 넘김) */
function applySidebarVisibility(activeGrps, activeCats){{
  document.querySelectorAll('.tree-group').forEach(el => {{
    el.style.display = activeGrps.has(el.dataset.grp) ? '' : 'none';
  }});
  document.querySelectorAll('.tree-cat-item').forEach(el => {{
    const key = el.dataset.grp + '|' + el.dataset.cat;
    el.style.display = activeCats.has(key) ? '' : 'none';
  }});
}}

/* 디바운스: 필터를 연속 클릭할 때 마지막 클릭 후 60ms 뒤에만 실제 render 실행 */
let _renderTimer = null;
function render(){{
  if(_renderTimer) return;   // 이미 예약되어 있으면 무시 (60ms 내 연속 호출 병합)
  _renderTimer = setTimeout(() => {{
    _renderTimer = null;
    _renderImpl();
  }}, 60);
}}
/* 즉시 실행 (초기 로드용) */
function renderNow(){{ if(_renderTimer){{ clearTimeout(_renderTimer); _renderTimer=null; }} _renderImpl(); }}

/* 최적화된 renderImpl(): RAW를 1회만 순회하여 사이드바/baseRows/rows/전체기간매출을 동시 구성 */
function _renderImpl(){{
  const unchecked = getSbUnchecked();
  const isYTD = S.months.has('YTD');
  const uncheckedSize = unchecked.size;
  const activeGrps = new Set(), activeCats = new Set();
  const baseRows = [], rows = [];
  const allT = S.teams.has('ALL'),   allBr = S.brands.has('ALL');
  const allCh = S.channels.has('ALL'), allCt = S.countries.has('ALL');
  const allCs = S.customers.has('ALL'), allTh = S.themes.has('ALL');
  const saleType = S.saleType;
  // 2026년 매출비중: 2026년 & 선택 월/판매구분에서의 매출액 기준
  let period26Rev = 0, filtered26Rev = 0;

  for(let i = 0; i < RAW.length; i++){{
    const d = RAW[i];
    const is26 = (String(d.yr) === '2026');
    // (A) 매출비중 분모: 2026년 + 팀/브랜드/채널/국가/거래처/테마 필터 무시 + 월/판매구분만 적용
    if(is26
       && (saleType === 'ALL' || d.sale_type === saleType)
       && (isYTD || S.months.has(d.month))){{
      period26Rev += d.rev;
    }}
    // 1. 팀/브랜드/채널/국가/거래처/테마
    if(!allT  && !S.teams.has(d.team))         continue;
    if(!allBr && !S.brands.has(d.brand))       continue;
    if(!allCh && !S.channels.has(d.channel))   continue;
    if(!allCt && !S.countries.has(d.country))  continue;
    if(!allCs && !S.customers.has(d.customer)) continue;
    if(!allTh && !S.themes.has(d.theme))       continue;
    if(d.item_group){{
      activeGrps.add(d.item_group);
      if(d.item_cat) activeCats.add(d.item_group+'|'+d.item_cat);
    }}
    // 2. 연도/판매구분/SKU
    if(!S.years.has(String(d.yr))) continue;
    if(saleType !== 'ALL' && d.sale_type !== saleType) continue;
    if(uncheckedSkus.has(d.sku)) continue;
    if(uncheckedSize && unchecked.has(d.item_cat)) continue;
    baseRows.push(d);
    // 3. 월/YTD 필터
    if(isYTD || S.months.has(d.month)){{
      rows.push(d);
      if(is26) filtered26Rev += d.rev;   // 2026년 매출비중 분자
    }}
  }}

  applySidebarVisibility(activeGrps, activeCats);
  renderKPI(rows);
  renderCharts(baseRows);
  renderTable(rows);
  // 2026년 매출비중 카드 갱신 (분자 = 필터 통과 2026 매출, 분모 = 2026 전체 기간 매출)
  const share = period26Rev ? (filtered26Rev / period26Rev * 100) : 0;
  const shareEl = document.getElementById('kRevShare');
  if(shareEl) shareEl.textContent = period26Rev ? share.toFixed(1) + '%' : '-';
  updateStickyOffsets();
  // 채널별 매출 현황 표(무거운 5레벨 표)는 브라우저 idle 시간에 렌더 → 반응성 대폭 개선
  //   다음 클릭이 빠르게 오면 이전 예약 취소 (계속 미뤄짐)
  if(_table2Cancel) _table2Cancel();
  const _hasIdle = (typeof requestIdleCallback === 'function');
  if(_hasIdle){{
    const id = requestIdleCallback(() => renderTable2(rows), {{timeout: 400}});
    _table2Cancel = () => cancelIdleCallback(id);
  }} else {{
    const id = setTimeout(() => renderTable2(rows), 60);
    _table2Cancel = () => clearTimeout(id);
  }}
}}
let _table2Cancel = null;

{init_js}
window.addEventListener('resize', () => {{
  updateStickyOffsets();
  if(_expandedChart) applyExpandRect(document.getElementById(_expandedChart.cardId));
}});

</script>
</body>
</html>"""
    return html


def main():
    import datetime, os
    print('=== 품목별 매출 대시보드 생성 ===')
    print('데이터 로드 중...')
    factory = load_factory()
    cmap, country_map, customer_map = load_channel_map()
    channel_order = load_channel_order()
    recs = []
    for yr in (2025, 2026):
        try:
            recs += load_sales(yr, factory, cmap, country_map, customer_map)
        except SystemExit as e:
            print(f'  [경고] {e} — 스킵')
    if not recs:
        sys.exit('[오류] 로드된 데이터가 없습니다.')
    print(f'  집계 중... 원시 {len(recs):,}건')
    records = aggregate(recs)
    print(f'  집계 완료: {len(records):,}건')
    chartjs_src = load_chartjs()
    base_date   = datetime.date.today().strftime('%Y년 %m월 %d일')
    html = make_html(records, chartjs_src, base_date, channel_order=channel_order)
    OUTPUT_HTML.write_text(html, encoding='utf-8')
    kb = OUTPUT_HTML.stat().st_size // 1024
    print(f'[완료] 생성: {OUTPUT_HTML}  ({kb} KB)')
    if chartjs_src:
        print('   Chart.js: 오프라인 내장')
    else:
        print('   Chart.js: CDN (인터넷 필요)')
    try:
        os.startfile(str(OUTPUT_HTML))
    except Exception:
        webbrowser.open(OUTPUT_HTML.as_uri())


if __name__ == '__main__':
    main()

